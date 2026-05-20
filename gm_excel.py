# ==============================================================
# EQEmu GM Command Report — Report Generator
# Project: https://github.com/YOUR_USERNAME/eqemu-gm-dashboard
# Version: 1.1  |  Created: May 2026
#
# Processes GM command TSV exports from an EQEmu/akk-stack
# server and produces a dark-mode HTML dashboard and Excel
# workbook with trends, item rarity, player impact, bots,
# online status, and new-command tracking.
#
# Free to use and adapt. Please keep this attribution intact.
# ==============================================================

import sys
import csv
import re
import json
import collections
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN_CMDS   = sys.argv[1]
IN_KILLS  = sys.argv[2]
IN_ITEMS  = sys.argv[3]
IN_IMPACT = sys.argv[4]
IN_BOTS   = sys.argv[5]
OUT_XLSX  = sys.argv[6]
OUT_HTML  = sys.argv[7] if len(sys.argv) > 7 else OUT_XLSX.replace(".xlsx", ".html")
NEW_COUNT = int(sys.argv[8]) if len(sys.argv) > 8 else -1
IN_CHARS  = sys.argv[9]  if len(sys.argv) > 9  else None
IN_DEATHS = sys.argv[10] if len(sys.argv) > 10 else None
# When passed, skip Excel generation entirely (saves ~10-15s per cycle).
HTML_ONLY  = "--html-only" in sys.argv

# Read last run timestamp from mounted state file
try:
    with open("/data/last_ts.txt") as f:
        LAST_TS = f.read().strip()
except Exception:
    LAST_TS = ""

CHARCOAL = "2D2D2D"
TEAL     = "1A7B72"
WHITE    = "FFFFFF"
ALT_ROW  = "F5F5F5"

H_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
H_FILL = PatternFill("solid", fgColor=CHARCOAL)
S_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
S_FILL = PatternFill("solid", fgColor=TEAL)
A_FILL = PatternFill("solid", fgColor=ALT_ROW)
BODY   = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
THIN   = Border(
    bottom=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin",  color="DDDDDD"),
)
PURPLE_FONT  = Font(name="Arial", size=10, color="7C3AED", bold=True)
AMBER_FONT   = Font(name="Arial", size=10, color="D97706", bold=True)
RED_FONT     = Font(name="Arial", size=10, color="DC2626", bold=True)

def excel_item_font(rarity):
    colors = {
        "orange": "FF8000",
        "purple": "A335EE",
        "blue":   "0070DD",
        "green":  "00AA00",
        "white":  "888888",
        "grey":   "666666",
    }
    hex_color = colors.get(rarity, "888888")
    bold = rarity in ("orange","purple","blue")
    return Font(name="Arial", size=10, color=hex_color, bold=bold)

def hrow(ws, cols, row=None):
    r = row or ws.max_row
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CENTER

def slabel(ws, label, ncols):
    ws.append([label] + [""] * (ncols - 1))
    r = ws.max_row
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = S_FONT; cell.fill = S_FILL
        cell.alignment = LEFT if c == 1 else CENTER

def drow(ws, values, alt=False):
    ws.append(values)
    r = ws.max_row
    for c in range(1, len(values) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BODY; cell.border = THIN
        if alt: cell.fill = A_FILL

def autofit(ws, mx=80):
    for col in ws.columns:
        w = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, mx)

def categorize(cmd):
    if not cmd: return "Unknown"
    c = cmd.lower()
    if "#giveitem" in c or "#summonitem" in c: return "Give Item"
    if "#finditem" in c:                        return "Find Item"
    if "#kill" in c:                            return "Kill"
    if "#level" in c:                           return "Set Level"
    if "#setaa" in c:                           return "Set AA"
    if "#buff" in c:                            return "Buff"
    if "#heal" in c:                            return "Heal"
    if "#mana" in c:                            return "Mana"
    if "#zone" in c:                            return "Zone"
    if "#goto" in c or "#moveto" in c:          return "Goto"
    if "#summon" in c:                          return "Summon"
    if "#god" in c or "#gm" in c:              return "GM Toggle"
    if "#spawn" in c:                           return "Spawn NPC"
    if "#npc" in c:                             return "NPC Edit"
    return "Other"

def cmd_html(cmd, category=''):
    """Return safe HTML for a command table cell.
    #giveitem/#summonitem → item ID links to lookupItem.
    #buff/#castspell etc. → spell ID links to lookupSpell.
    Buff category fallback: first number treated as spell ID."""
    if not cmd:
        return ''
    stripped = cmd.strip()
    # Give Item / Summon Item → item ID clickable via lookupItem
    im = ITEM_RE.match(stripped)
    if im:
        item_id = im.group(1)
        prefix  = esc(stripped[:im.start(1)])
        tail    = esc(stripped[im.end():])
        return (f'{prefix}<span class="item-link" onclick="lookupItem({item_id})" '
                f'title="Item #{item_id}">{item_id}</span>{tail}')
    # Spell commands → spell ID clickable via lookupSpell
    m = BUFF_RE.match(stripped)
    if m:
        verb     = esc(m.group(1))
        spell_id = m.group(2)
        tail     = esc(m.group(3) or '')
        return (f'{verb} <span class="spell-link" onclick="lookupSpell({spell_id})" '
                f'title="Look up spell #{spell_id}">{spell_id}</span>{tail}')
    # Fallback: Buff category → first number is likely spell ID
    if category == 'Buff':
        nm = re.search(r'(\d+)', stripped)
        if nm:
            spell_id = nm.group(1)
            before   = esc(stripped[:nm.start()])
            after    = esc(stripped[nm.end():])
            return (f'{before}<span class="spell-link" onclick="lookupSpell({spell_id})" '
                    f'title="Look up spell #{spell_id}">{spell_id}</span>{after}')
    return esc(cmd)


ACCOUNT_SUFFIXES = sorted(
    ['admin','administrator','hill','user','users','test','dev','ops',
     'sys','net','web','mod','gm','mgr','mgmt','acct','guest'],
    key=len, reverse=True
)

def account_initials(account, cmd_count=None, threshold=10):
    """Convert account name to dot-separated initials.
    Returns empty string if cmd_count is below threshold."""
    if cmd_count is not None and cmd_count < threshold:
        return ""
    name = account.lower()
    for suffix in ACCOUNT_SUFFIXES:
        if name.endswith(suffix) and len(name) > len(suffix):
            name = name[:-len(suffix)]
            break
    part = name[:3] if name else account[:2]
    return ".".join(list(part)) + "." if part else account[:1] + "."



ITEM_RE  = re.compile(r'^#(?:giveitem|summonitem)\s+(\d+)(?:\s+(\d+))?', re.IGNORECASE)
BUFF_RE  = re.compile(r'^(#(?:buff|castspell|discipline))\s+(\d+)(.*)?$', re.IGNORECASE)

def parse_item_command(raw_cmd, sql_item_id):
    m = ITEM_RE.match(raw_cmd.strip()) if raw_cmd else None
    if m:
        return m.group(1), int(m.group(2)) if m.group(2) else 1
    return sql_item_id, 1

GENERIC_ARTICLE = re.compile(r'^(a_|an_|the_)', re.IGNORECASE)
GENERIC_SUFFIX  = re.compile(
    r'_(goblin|orc|gnoll|kobold|elemental|guard|warrior|knight|soldier|'
    r'champion|scout|hunter|skeleton|zombie|ghoul|ghost|spectre|specter|'
    r'wraith|revenant|mummy|bat|rat|wolf|bear|spider|snake|beetle|worm|'
    r'golem|gargoyle|bandit|brigand|thug|assassin|ritualist|mystic|'
    r'shaman|berserker|raider|fiend|demon|imp|drake|wyvern|lizard|cat|'
    r'gnome|dwarf|elf|troll|ogre|halfling|barbarian|iksar|sarnak|'
    r'mercenary|wisp|treant|dryad|fairy|pixie|flesh|bone|shade|'
    r'spirit|froglok|brownie|centaur|minotaur|harpy|siren|mermaid|'
    r'aviak|griffon|sphinx|chimera|basilisk|manticore)\d*$',
    re.IGNORECASE
)

def is_named_mob(name):
    if not name or not name[0].isupper(): return False
    base = re.sub(r'\d+$', '', name).rstrip('_')
    if GENERIC_ARTICLE.match(base): return False
    if GENERIC_SUFFIX.search(base): return False
    return True

def kill_target_class(name):
    return "target named-target" if is_named_mob(name) else "target"

def detail_class(cat, detail, cmd=""):
    if cat == "Kill" and is_named_mob(detail):
        return "named-target"
    if cat == "Give Item":
        m = ITEM_RE.match(cmd.strip()) if cmd else None
        if m:
            rarity = item_rarity_lookup.get(m.group(1), "grey")
            return item_class(rarity)
        return "item-grey"
    return "item-name"

EQ_ITEM_COLORS = {
    "orange": "#ff8000",
    "purple": "#a335ee",
    "blue":   "#0070dd",
    "green":  "#1eff00",
    "white":  "#d0d0d0",
    "grey":   "#9d9d9d",
}

def classify_item(row):
    """Return EQ-style item color tier based on item properties."""
    try:
        nodrop   = int(row.get('nodrop',   0) or 0)
        reqlevel = int(row.get('reqlevel', 0) or 0)
        hp       = int(row.get('hp',       0) or 0)
        mana     = int(row.get('mana',     0) or 0)
        ac       = int(row.get('ac',       0) or 0)
        damage   = int(row.get('damage',   0) or 0)
        magic    = int(row.get('magic',    0) or 0)
        price    = int(row.get('price',    0) or 0)
    except (ValueError, TypeError):
        return 'grey'

    stat_score = hp + mana + (ac * 8) + (damage * 15)

    if nodrop and reqlevel >= 60 and stat_score >= 600:
        return 'orange'
    if nodrop and (reqlevel >= 46 or stat_score >= 350):
        return 'purple'
    if nodrop or reqlevel >= 35 or stat_score >= 175:
        return 'blue'
    if magic and (reqlevel >= 10 or stat_score >= 40):
        return 'green'
    if magic or price >= 1000:
        return 'white'
    return 'grey'

def load_tsv(path, fields):
    rows = []
    try:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f, delimiter="\t", fieldnames=fields):
                rows.append(row)
    except Exception:
        pass
    return rows

cmds   = load_tsv(IN_CMDS,   ["timestamp","character_name","account_name","account_status","command"])
kills  = load_tsv(IN_KILLS,  ["timestamp","character_name","target_name","target_id"])
items  = load_tsv(IN_ITEMS,  ["timestamp","character_name","item_id","item_name","nodrop","reqlevel","hp","mana","ac","damage","magic","lore","price","raw_command"])
impact = load_tsv(IN_IMPACT, ["timestamp","gm_character","recipient","item_id","item_name","nodrop","reqlevel","magic","lore","price","raw_command"])
bots   = load_tsv(IN_BOTS,   ["account_name","owner_name","bot_name","bot_class","bot_race","bot_gender","bot_level","bot_hp","bot_mana"])

EQ_CLASSES = {1:"Warrior",2:"Cleric",3:"Paladin",4:"Ranger",5:"Shadow Knight",
              6:"Druid",7:"Monk",8:"Bard",9:"Rogue",10:"Shaman",
              11:"Necromancer",12:"Wizard",13:"Magician",14:"Enchanter",
              15:"Beastlord",16:"Berserker"}
EQ_RACES   = {1:"Human",2:"Barbarian",3:"Erudite",4:"Wood Elf",5:"High Elf",
              6:"Dark Elf",7:"Half Elf",8:"Dwarf",9:"Troll",10:"Ogre",
              11:"Halfling",12:"Gnome",13:"Iksar",14:"Vahshir",15:"Froglok",128:"Drakkin"}

chars_raw  = load_tsv(IN_CHARS,  ["name","class","level","race","account_name","account_status","guild_name","last_login"]) if IN_CHARS  else []
deaths_raw = load_tsv(IN_DEATHS, ["timestamp","character_name","level","killed_by","zone","spell_id"])              if IN_DEATHS else []
for r in chars_raw:
    r["class_name"] = EQ_CLASSES.get(int(r.get("class",0) or 0), f"Class {r.get('class','?')}")
    r["race_name"]  = EQ_RACES.get(int(r.get("race",0) or 0),   f"Race {r.get('race','?')}")
    r["is_gm"]      = int(r.get("account_status",0) or 0) >= 100
online = []
try:
    online = load_tsv("/data/eqemu_gm_online.tsv", ["character_name","account_name","zone_name","level","class_id","seconds_online"])
except Exception:
    pass

offline = []
try:
    offline = load_tsv("/data/eqemu_gm_offline.tsv", ["character_name","account_name","level","class_id","last_login_time","seconds_offline","is_gm"])
except Exception:
    pass

EQ_CLASSES = {
    "1":"Warrior","2":"Cleric","3":"Paladin","4":"Ranger","5":"Shadow Knight",
    "6":"Druid","7":"Monk","8":"Bard","9":"Rogue","10":"Shaman",
    "11":"Necromancer","12":"Wizard","13":"Magician","14":"Enchanter",
    "15":"Beastlord","16":"Berserker"
}
EQ_RACES = {
    "1":"Human","2":"Barbarian","3":"Erudite","4":"Wood Elf","5":"High Elf",
    "6":"Dark Elf","7":"Half Elf","8":"Dwarf","9":"Troll","10":"Ogre",
    "11":"Halfling","12":"Gnome","13":"Iksar","14":"Vah Shir","15":"Froglok","16":"Drakkin"
}
EQ_GENDERS = {"0":"Male","1":"Female","2":"Neuter"}

for r in bots:
    r["class_name"]  = EQ_CLASSES.get(r.get("bot_class",""),  r.get("bot_class","?"))
    r["race_name"]   = EQ_RACES.get(r.get("bot_race",""),     r.get("bot_race","?"))
    r["gender_name"] = EQ_GENDERS.get(r.get("bot_gender",""), r.get("bot_gender","?"))

for r in online:
    r["class_name"] = EQ_CLASSES.get(r.get("class_id",""), "Unknown")

for r in offline:
    r["class_name"] = EQ_CLASSES.get(r.get("class_id",""), "Unknown")

def format_duration(seconds):
    try:
        s = int(float(seconds))
    except (ValueError, TypeError):
        return "?"
    if s < 60:    return f"{s}s"
    if s < 3600:  return f"{s//60}m"
    if s < 86400:
        h = s // 3600; m = (s % 3600) // 60
        return f"{h}h {m}m"
    d = s // 86400; h = (s % 86400) // 3600
    return f"{d}d {h}h"

# Parse items
for r in items:
    parsed_id, qty = parse_item_command(r.get("raw_command",""), r.get("item_id",""))
    r["item_id"]  = parsed_id
    r["quantity"] = qty
    r["rarity"]   = classify_item(r)
    if r["item_name"] in ("Unknown Item","",None) and parsed_id:
        r["item_name"] = f"Unknown (ID: {parsed_id})"

# Parse impact
for r in impact:
    parsed_id, qty = parse_item_command(r.get("raw_command",""), r.get("item_id",""))
    r["item_id"]  = parsed_id
    r["quantity"] = qty
    r["rarity"]   = classify_item(r)
    if r["item_name"] in ("Unknown Item","",None) and parsed_id:
        r["item_name"] = f"Unknown (ID: {parsed_id})"

item_lookup = {}
item_rarity_lookup = {}
for r in items:
    if r["item_id"]:
        if r["item_name"] and not r["item_name"].startswith("Unknown"):
            item_lookup[r["item_id"]] = r["item_name"]
        elif r["item_id"] not in item_lookup:
            item_lookup[r["item_id"]] = r["item_name"]
        item_rarity_lookup[r["item_id"]] = r["rarity"]

# Reverse lookup: item_name → item_id (for places that only have the name)
name_to_id = {v: k for k, v in item_lookup.items()}

kill_lookup = {}
for r in kills:
    kill_lookup[(r["timestamp"], r["character_name"])] = r["target_name"]

def resolve_detail(row):
    cmd = row.get("command","")
    cat = row.get("category","")
    if cat == "Give Item":
        m = ITEM_RE.match(cmd.strip()) if cmd else None
        if m:
            return item_lookup.get(m.group(1), f"Unknown (ID: {m.group(1)})")
    if cat == "Kill":
        target = kill_lookup.get((row.get("timestamp",""), row.get("character_name","")), "")
        return target if target and target != "Unknown" else ""
    return ""

def give_item_id(r):
    """Return item_id string for a Give Item command row, or empty string."""
    m = ITEM_RE.match(r.get("command","").strip()) if r.get("category") == "Give Item" else None
    return m.group(1) if m else ""

def detail_html(r):
    """Return HTML for the Item/Target detail cell.
    Give Item rows get a clickable item name; all others get plain escaped text."""
    detail = r.get("detail","")
    if not detail:
        return ""
    iid = give_item_id(r)
    if iid:
        rarity   = item_rarity_lookup.get(iid, "grey")
        css      = item_class(rarity)
        safe_nm  = esc(detail)
        return (f'<span class="{css} item-link" ' 
                f'onclick="lookupItem({iid})" ' 
                f'style="cursor:pointer;text-decoration:underline dotted">' 
                f'{safe_nm}</span>')
    if r.get("category") == "Kill":
        return f'<span class="{detail_class(r["category"],detail,r.get("command",""))}">{esc(detail)}</span>'
    return esc(detail)

for r in cmds:
    r["category"] = categorize(r.get("command",""))
    r["detail"]   = resolve_detail(r)

# ---- Aggregations ----
total          = len(cmds)
by_char        = collections.Counter(r["character_name"] for r in cmds)
by_cat         = collections.Counter(r["category"]       for r in cmds)
kill_by_target = collections.Counter(r["target_name"] for r in kills if r["target_name"] != "Unknown")
item_by_name   = collections.Counter(r["item_name"]   for r in items)
all_categories = [cat for cat, _ in by_cat.most_common()]
generated      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# GM = account.status >= 100 — not just anyone who typed a # command
gm_names = set(
    r["character_name"] for r in cmds
    if (lambda v: int(v) >= 100 if str(v or 0).lstrip('-').isdigit() else False)(r.get("account_status", 0) or 0)
)

def gm_name(name):
    safe = esc(name).replace("'", "&#39;")
    if name in gm_names:
        return f'<span style="color:#fb923c;font-weight:700;cursor:pointer" onclick="lookupInventory(\'{safe}\',\'character\')">{esc(name)}</span>'
    return f'<span style="cursor:pointer;text-decoration:underline dotted;color:var(--text2)" onclick="lookupInventory(\'{safe}\',\'character\')">{esc(name)}</span>'

# Pre-compute display initials per character (hidden if < 10 commands)
char_initials = {
    ch: account_initials(
        next((r["account_name"] for r in cmds if r["character_name"]==ch), ""),
        cnt
    )
    for ch, cnt in by_char.items()
}

# Trend data
daily_counts  = collections.Counter()
hourly_counts = collections.Counter()
char_daily    = collections.defaultdict(lambda: collections.Counter())
for r in cmds:
    try:
        dt = datetime.strptime(r["timestamp"], "%Y-%m-%d %H:%M:%S")
        day = dt.strftime("%Y-%m-%d")
        daily_counts[day] += 1
        hourly_counts[dt.hour] += 1
        char_daily[r["character_name"]][day] += 1
    except Exception:
        pass

sorted_days  = sorted(daily_counts.keys())
hourly_data  = [hourly_counts.get(h, 0) for h in range(24)]

# Day-of-week and peak hour breakdowns
DOW_NAMES = ["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
dow_counts = collections.Counter()
for r in cmds:
    try:
        dt = datetime.strptime(r["timestamp"], "%Y-%m-%d %H:%M:%S")
        dow_counts[dt.weekday()] += 1   # 0=Monday in Python
    except Exception:
        pass

# Peak hours ranked
peak_hours = sorted(
    [(h, hourly_counts.get(h, 0)) for h in range(24)],
    key=lambda x: -x[1]
)

# Time-of-day buckets
def time_bucket(hour):
    if   0  <= hour < 6:  return "Late Night (12am-6am)"
    elif 6  <= hour < 12: return "Morning (6am-12pm)"
    elif 12 <= hour < 18: return "Afternoon (12pm-6pm)"
    else:                 return "Evening (6pm-12am)"

bucket_counts = collections.Counter()
for h, cnt in hourly_counts.items():
    bucket_counts[time_bucket(h)] += cnt

# Impact aggregations
impact_by_recipient = collections.Counter(r["recipient"] for r in impact if r["recipient"] != "Unknown")
impact_by_gm        = collections.Counter(r["gm_character"] for r in impact)
bots_by_account = collections.defaultdict(list)
for r in bots:
    bots_by_account[r["account_name"]].append(r)
bots_by_class   = collections.Counter(r["class_name"] for r in bots)
total_bots      = len(bots)

new_label = str(NEW_COUNT) if NEW_COUNT >= 0 else "0"
delta_border = ".5px solid var(--border)"

# Excel uses server-side new_cmds; web uses client-side JS
if LAST_TS:
    try:
        last_dt  = datetime.strptime(LAST_TS.strip(), "%Y-%m-%d %H:%M:%S")
        new_cmds = [r for r in cmds if datetime.strptime(r["timestamp"], "%Y-%m-%d %H:%M:%S") > last_dt]
    except Exception:
        new_cmds = []
else:
    new_cmds = []

if not HTML_ONLY:
    # ============================================================
    # EXCEL
    # ============================================================
    wb = Workbook()
    
    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "GM Command Usage"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=CHARCOAL)
    ws["B1"] = f"Generated: {generated}   |   Total: {total}   |   New since last report: {new_label}"
    ws["B1"].font = Font(name="Arial", size=10, color="666666")
    ws.append([])
    slabel(ws, "Commands by Character", 4); hrow(ws, ["Character","Account","Total","% of Total"])
    for i, (ch, cnt) in enumerate(sorted(by_char.items(), key=lambda x: -x[1])):
        if cnt < 10:
            continue
        acct = char_initials.get(ch, "")
        ws.append([ch, acct, cnt, f"{cnt/total*100:.1f}%"])
        rn = ws.max_row
        for c in range(1,5):
            ws.cell(row=rn,column=c).font=BODY; ws.cell(row=rn,column=c).border=THIN
            if i%2==0: ws.cell(row=rn,column=c).fill=A_FILL
    ws.append([])
    slabel(ws, "Commands by Type", 3); hrow(ws, ["Type","Count","% of Total"])
    for i, (cat, cnt) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1])):
        ws.append([cat, cnt, f"{cnt/total*100:.1f}%"])
        rn = ws.max_row
        for c in range(1,4):
            ws.cell(row=rn,column=c).font=BODY; ws.cell(row=rn,column=c).border=THIN
            if i%2==0: ws.cell(row=rn,column=c).fill=A_FILL
    autofit(ws)
    
    # All Commands
    ws_all = wb.create_sheet("All Commands")
    ws_all.sheet_view.showGridLines = False; ws_all.freeze_panes = "A2"
    hrow(ws_all, ["Timestamp","Character","Account","Category","Command","Item / Target"], row=1)
    for i, r in enumerate(cmds):
        drow(ws_all,[r["timestamp"],r["character_name"],char_initials.get(r["character_name"],""),r["category"],r["command"],r["detail"]],alt=(i%2==0))
    ws_all.column_dimensions["A"].width=22; ws_all.column_dimensions["B"].width=16
    ws_all.column_dimensions["C"].width=16; ws_all.column_dimensions["D"].width=14
    ws_all.column_dimensions["E"].width=28; ws_all.column_dimensions["F"].width=40
    
    # Trends
    ws_tr = wb.create_sheet("Trends")
    ws_tr.sheet_view.showGridLines = False
    ws_tr["A1"] = "Command Trends"
    ws_tr["A1"].font = Font(name="Arial", bold=True, size=14, color=CHARCOAL)
    ws_tr.append([])
    slabel(ws_tr, "Daily Command Counts", 2); hrow(ws_tr, ["Date","Count"])
    for i, day in enumerate(sorted_days):
        drow(ws_tr, [day, daily_counts[day]], alt=(i%2==0))
    ws_tr.append([])
    slabel(ws_tr, "Commands by Hour of Day", 2); hrow(ws_tr, ["Hour","Count"])
    for h in range(24):
        drow(ws_tr, [f"{h:02d}:00", hourly_counts.get(h,0)], alt=(h%2==0))
    ws_tr.append([])
    slabel(ws_tr, "Peak Hours Ranked", 2); hrow(ws_tr, ["Hour","Commands"])
    for i, (h, cnt) in enumerate(peak_hours):
        ampm = f"{h % 12 or 12}{'am' if h < 12 else 'pm'}"
        drow(ws_tr, [f"{h:02d}:00  ({ampm})", cnt], alt=(i%2==0))
    ws_tr.append([])
    slabel(ws_tr, "Commands by Time of Day", 2); hrow(ws_tr, ["Period","Commands"])
    for i, bucket in enumerate(["Morning (6am-12pm)","Afternoon (12pm-6pm)","Evening (6pm-12am)","Late Night (12am-6am)"]):
        drow(ws_tr, [bucket, bucket_counts.get(bucket, 0)], alt=(i%2==0))
    ws_tr.append([])
    slabel(ws_tr, "Commands by Day of Week", 2); hrow(ws_tr, ["Day","Commands"])
    py_to_eq = {0:"Monday",1:"Tuesday",2:"Wednesday",3:"Thursday",4:"Friday",5:"Saturday",6:"Sunday"}
    for i, (dow_idx, name) in enumerate(sorted(py_to_eq.items())):
        drow(ws_tr, [name, dow_counts.get(dow_idx, 0)], alt=(i%2==0))
    ws_tr.column_dimensions["A"].width=20; ws_tr.column_dimensions["B"].width=14
    
    # Player Impact
    ws_pi = wb.create_sheet("Player Impact")
    ws_pi.sheet_view.showGridLines = False
    ws_pi["A1"] = "Player Impact — Items Given by Target"
    ws_pi["A1"].font = Font(name="Arial", bold=True, size=14, color=CHARCOAL)
    ws_pi.append([])
    slabel(ws_pi, "Items Received by Character", 3); hrow(ws_pi, ["Character","Items Received",""])
    for i, (char, cnt) in enumerate(impact_by_recipient.most_common()):
        drow(ws_pi, [char, cnt, ""], alt=(i%2==0))
    ws_pi.append([])
    slabel(ws_pi, "Full Impact Log", 6)
    hrow(ws_pi, ["Timestamp","GM Character","Recipient","Item ID","Item Name","Rarity"])
    for i, r in enumerate(impact):
        drow(ws_pi,[r["timestamp"],r["gm_character"],r["recipient"],r["item_id"],r["item_name"],r["rarity"].title()],alt=(i%2==0))
        rn = ws_pi.max_row
        ws_pi.cell(row=rn, column=5).font = excel_item_font(r["rarity"])
        ws_pi.cell(row=rn, column=6).font = excel_item_font(r["rarity"])
    ws_pi.column_dimensions["A"].width=22; ws_pi.column_dimensions["B"].width=18
    ws_pi.column_dimensions["C"].width=18; ws_pi.column_dimensions["D"].width=12
    ws_pi.column_dimensions["E"].width=40; ws_pi.column_dimensions["F"].width=12
    
    # Activity Details
    ws_det = wb.create_sheet("Activity Details")
    ws_det.sheet_view.showGridLines = False
    ws_det["A1"] = "Activity Details — Kills & Items Given"
    ws_det["A1"].font = Font(name="Arial", bold=True, size=14, color=CHARCOAL)
    ws_det.append([])
    slabel(ws_det, f"Mobs Killed ({len(kills)} events)", 4)
    hrow(ws_det, ["Timestamp","GM Character","Target Name","Target ID"])
    for i, r in enumerate(kills):
        drow(ws_det,[r["timestamp"],r["character_name"],r["target_name"],r["target_id"]],alt=(i%2==0))
        if is_named_mob(r["target_name"]):
            ws_det.cell(row=ws_det.max_row, column=3).font = PURPLE_FONT
    ws_det.append([])
    slabel(ws_det, "Kill Summary", 3); hrow(ws_det, ["Target","Kill Count",""])
    for i, (t, cnt) in enumerate(kill_by_target.most_common()):
        drow(ws_det,[t, cnt, ""],alt=(i%2==0))
        if is_named_mob(t):
            ws_det.cell(row=ws_det.max_row, column=1).font = PURPLE_FONT
    ws_det.append([])
    slabel(ws_det, f"Items Given ({len(items)} events)", 7)
    hrow(ws_det, ["Timestamp","GM Character","Item ID","Item Name","Rarity","Qty","Command"])
    for i, r in enumerate(items):
        drow(ws_det,[r["timestamp"],r["character_name"],r["item_id"],r["item_name"],r["rarity"].title(),r["quantity"],r["raw_command"]],alt=(i%2==0))
        rn = ws_det.max_row
        ws_det.cell(row=rn, column=4).font = excel_item_font(r["rarity"])
        ws_det.cell(row=rn, column=5).font = excel_item_font(r["rarity"])
    ws_det.append([])
    slabel(ws_det, "Item Summary", 3); hrow(ws_det, ["Item Name","Times Given",""])
    for i, (nm, cnt) in enumerate(item_by_name.most_common()):
        drow(ws_det,[nm, cnt, ""],alt=(i%2==0))
    ws_det.column_dimensions["A"].width=22; ws_det.column_dimensions["B"].width=18
    ws_det.column_dimensions["C"].width=12; ws_det.column_dimensions["D"].width=40
    ws_det.column_dimensions["E"].width=12; ws_det.column_dimensions["F"].width=6
    ws_det.column_dimensions["G"].width=28
    
    # Bots
    ws_bots = wb.create_sheet("Bots")
    ws_bots.sheet_view.showGridLines = False
    ws_bots["A1"] = "Bots by Account"
    ws_bots["A1"].font = Font(name="Arial", bold=True, size=14, color=CHARCOAL)
    ws_bots["B1"] = f"Total bots: {total_bots}"
    ws_bots["B1"].font = Font(name="Arial", size=10, color="666666")
    ws_bots.append([])
    for acct in sorted(bots_by_account.keys()):
        acct_bots = bots_by_account[acct]
        slabel(ws_bots, f"{account_initials(acct, 999)}  ({acct})  —  {len(acct_bots)} bots", 8)
        hrow(ws_bots, ["Character","Bot Name","Class","Race","Gender","Level","HP","Mana"])
        for i, r in enumerate(sorted(acct_bots, key=lambda x: (x["owner_name"], x["bot_name"]))):
            drow(ws_bots,[
                r["owner_name"], r["bot_name"], r["class_name"], r["race_name"],
                r["gender_name"], r.get("bot_level",""),
                r.get("bot_hp",""), r.get("bot_mana","")
            ], alt=(i%2==0))
        ws_bots.append([])
    ws_bots.column_dimensions["A"].width=18; ws_bots.column_dimensions["B"].width=22
    ws_bots.column_dimensions["C"].width=16; ws_bots.column_dimensions["D"].width=14
    ws_bots.column_dimensions["E"].width=10; ws_bots.column_dimensions["F"].width=8
    ws_bots.column_dimensions["G"].width=10; ws_bots.column_dimensions["H"].width=10
    
    # Per command type tabs
    for cat in all_categories:
        cat_rows   = [r for r in cmds if r["category"] == cat]
        has_detail = cat in ("Give Item","Kill")
        wt = wb.create_sheet(cat[:31])
        wt.sheet_view.showGridLines = False; wt.freeze_panes = "A3"
        wt["A1"] = f"Command Type: {cat}"
        wt["A1"].font = Font(name="Arial", bold=True, size=13, color=CHARCOAL)
        wt["B1"] = f"Total: {len(cat_rows)}"
        wt["B1"].font = Font(name="Arial", size=10, color="666666")
        wt.append([])
        slabel(wt, "By Character", 3); hrow(wt, ["Character","Count","%"])
        cb = collections.Counter(r["character_name"] for r in cat_rows)
        for i, (ch, cnt) in enumerate(sorted(cb.items(), key=lambda x: -x[1])):
            wt.append([ch, cnt, f"{cnt/len(cat_rows)*100:.1f}%"])
            rn=wt.max_row
            for c in range(1,4):
                wt.cell(row=rn,column=c).font=BODY; wt.cell(row=rn,column=c).border=THIN
                if i%2==0: wt.cell(row=rn,column=c).fill=A_FILL
        wt.append([])
        if has_detail:
            slabel(wt, "Command Log", 5); hrow(wt, ["Timestamp","Character","Account","Command","Item / Target"])
            for i, r in enumerate(cat_rows):
                drow(wt,[r["timestamp"],r["character_name"],char_initials.get(r["character_name"],""),r["command"],r["detail"]],alt=(i%2==0))
                if cat == "Kill" and is_named_mob(r["detail"]):
                    wt.cell(row=wt.max_row, column=5).font = PURPLE_FONT
            wt.column_dimensions["E"].width=40
        else:
            slabel(wt, "Command Log", 4); hrow(wt, ["Timestamp","Character","Account","Command"])
            for i, r in enumerate(cat_rows):
                drow(wt,[r["timestamp"],r["character_name"],char_initials.get(r["character_name"],""),r["command"]],alt=(i%2==0))
        wt.column_dimensions["A"].width=22; wt.column_dimensions["B"].width=18
        wt.column_dimensions["C"].width=18; wt.column_dimensions["D"].width=28
    
    # Per-character tabs
    for char in sorted(by_char.keys()):
        cr = [r for r in cmds if r["character_name"]==char]
        wc = wb.create_sheet(char[:31])
        wc.sheet_view.showGridLines = False; wc.freeze_panes = "A3"
        wc["A1"] = f"GM Commands — {char}"
        wc["A1"].font = Font(name="Arial", bold=True, size=13, color=CHARCOAL)
        wc["B1"] = f"Total: {len(cr)}"
        wc["B1"].font = Font(name="Arial", size=10, color="666666")
        wc.append([])
        slabel(wc,"Breakdown by Type",3); hrow(wc,["Type","Count","%"])
        cc = collections.Counter(r["category"] for r in cr)
        for i,(cat,cnt) in enumerate(sorted(cc.items(),key=lambda x:-x[1])):
            wc.append([cat,cnt,f"{cnt/len(cr)*100:.1f}%"])
            rn=wc.max_row
            for c in range(1,4):
                wc.cell(row=rn,column=c).font=BODY; wc.cell(row=rn,column=c).border=THIN
                if i%2==0: wc.cell(row=rn,column=c).fill=A_FILL
        wc.append([])
        slabel(wc,"Command Log",5); hrow(wc,["Timestamp","Category","Command","Item / Target",""])
        for i,r in enumerate(cr):
            drow(wc,[r["timestamp"],r["category"],r["command"],r["detail"],""],alt=(i%2==0))
            if r["category"] == "Kill" and is_named_mob(r["detail"]):
                wc.cell(row=wc.max_row, column=4).font = PURPLE_FONT
        wc.column_dimensions["A"].width=22; wc.column_dimensions["B"].width=14
        wc.column_dimensions["C"].width=28; wc.column_dimensions["D"].width=40
    
    if new_cmds:
        ws_new = wb.create_sheet("New Commands")
        ws_new.sheet_view.showGridLines = False
        ws_new.freeze_panes = "A2"
        ws_new["A1"] = f"New Commands since {LAST_TS}"
        ws_new["A1"].font = Font(name="Arial", bold=True, size=13, color=CHARCOAL)
        ws_new.append([])
        hrow(ws_new, ["Timestamp","Character","Account","Category","Command","Item / Target"], row=2)
        for i, r in enumerate(new_cmds):
            drow(ws_new,[r["timestamp"],r["character_name"],char_initials.get(r["character_name"],""),r["category"],r["command"],r["detail"]],alt=(i%2==0))
        ws_new.column_dimensions["A"].width=22; ws_new.column_dimensions["B"].width=16
        ws_new.column_dimensions["C"].width=10; ws_new.column_dimensions["D"].width=14
        ws_new.column_dimensions["E"].width=28; ws_new.column_dimensions["F"].width=40
    
    wb.save(OUT_XLSX)


# ============================================================
# HTML
# ============================================================
def esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

def badge(cat):
    slug = cat.lower().replace(" ","-")
    return f'<span class="badge cat-{esc(slug)}">{esc(cat)}</span>'

def rarity_badge(r):
    color = EQ_ITEM_COLORS.get(r, "#9d9d9d")
    labels = {"orange":"Orange","purple":"Purple","blue":"Blue","green":"Green","white":"White","grey":"Grey"}
    label  = labels.get(r, r.title())
    return f'<span class="badge" style="background:{color}22;color:{color};border:1px solid {color}55">{label}</span>'

def item_color(rarity):
    return EQ_ITEM_COLORS.get(rarity, "#9d9d9d")

def item_class(rarity):
    return f"item-{rarity}"

chars = sorted(c for c in by_char.keys() if c in gm_names)

char_tab_btns = "".join(
    f'<button class="tab-btn filter-btn" onclick="showTab(\'char-{c.replace(" ","_")}\',this)" style="color:#fb923c">{esc(c)}</button>'
    if c in gm_names else
    f'<button class="tab-btn filter-btn" onclick="showTab(\'char-{c.replace(" ","_")}\',this)">{esc(c)}</button>'
    for c in chars
)
cat_tab_btns = "".join(
    f'<button class="tab-btn filter-btn" onclick="showTab(\'cat-{cat.lower().replace(" ","-")}\',this)">{badge(cat)} {by_cat[cat]}</button>'
    for cat in all_categories
)

# Trend JSON for Chart.js
trend_labels  = json.dumps(sorted_days[-90:] if len(sorted_days) > 90 else sorted_days)
trend_data    = json.dumps([daily_counts[d] for d in (sorted_days[-90:] if len(sorted_days) > 90 else sorted_days)])
hourly_labels = json.dumps([f"{h:02d}:00" for h in range(24)])
hourly_json   = json.dumps(hourly_data)
cat_labels    = json.dumps([c for c, _ in by_cat.most_common()])
cat_data      = json.dumps([cnt for _, cnt in by_cat.most_common()])

dow_order  = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
dow_map    = {0:"Monday",1:"Tuesday",2:"Wednesday",3:"Thursday",4:"Friday",5:"Saturday",6:"Sunday"}
dow_labels = json.dumps(dow_order)
dow_data   = json.dumps([dow_counts.get(i, 0) for i in range(7)])

peak_hours_html = "".join(
    f"<tr><td style='font-family:monospace'>{h:02d}:00</td>"
    f"<td style='color:#888;font-size:.8rem'>{h % 12 or 12}{'am' if h < 12 else 'pm'}</td>"
    f"<td style='font-weight:600'>{cnt}</td>"
    f"<td style='width:120px'><div style='height:8px;border-radius:4px;background:#1a7b72;width:{int(cnt/max(c for _,c in peak_hours)*100) if peak_hours and max(c for _,c in peak_hours)>0 else 0}%'></div></td></tr>"
    for h, cnt in peak_hours if cnt > 0
)

bucket_order = ["Morning (6am-12pm)","Afternoon (12pm-6pm)","Evening (6pm-12am)","Late Night (12am-6am)"]
bucket_icons = {"Morning (6am-12pm)":"🌅","Afternoon (12pm-6pm)":"☀️","Evening (6pm-12am)":"🌆","Late Night (12am-6am)":"🌙"}
bucket_html  = "".join(
    f"<tr><td>{bucket_icons.get(b,'')} {esc(b)}</td><td style='font-weight:600'>{bucket_counts.get(b,0)}</td>"
    f"<td style='width:150px'><div style='height:8px;border-radius:4px;background:#1a7b72;width:{int(bucket_counts.get(b,0)/max(bucket_counts.values())*100) if bucket_counts else 0}%'></div></td></tr>"
    for b in bucket_order
)

dow_html = "".join(
    f"<tr><td>{esc(dow_map[i])}</td><td style='font-weight:600'>{dow_counts.get(i,0)}</td>"
    f"<td style='width:150px'><div style='height:8px;border-radius:4px;background:#1a7b72;width:{int(dow_counts.get(i,0)/max(dow_counts.values())*100) if dow_counts else 0}%'></div></td></tr>"
    for i in range(7)
)

# Player impact HTML
impact_by_recipient_html = "".join(
    f"<tr><td class='item-name'>{esc(ch)}</td><td>{cnt}</td></tr>"
    for ch,cnt in impact_by_recipient.most_common()
) or '<tr><td colspan="2" style="color:#888;text-align:center">No target data available — check event_data logging config</td></tr>'

impact_log_html = "".join(
    f"<tr><td>{esc(r['timestamp'])}</td><td>{gm_name(r['gm_character'])}</td>"
    f"<td class='item-name'>{esc(r['recipient'])}</td>"
    f"<td>{esc(r['item_id'])}</td>"
    "<td class='" + item_class(r['rarity']) + "' style='cursor:pointer;text-decoration:underline dotted' onclick='lookupItem(" + r['item_id'] + ")'>" + esc(r['item_name']) + "</td>"
    f"<td>{rarity_badge(r['rarity'])}</td></tr>"
    for r in impact
) or '<tr><td colspan="6" style="color:#888;text-align:center">No impact data</td></tr>'

# Cat panes
cat_panes = ""
for cat in all_categories:
    cat_rows   = [r for r in cmds if r["category"] == cat]
    cid        = cat.lower().replace(" ","-")
    has_detail = cat in ("Give Item","Kill")
    cb         = collections.Counter(r["character_name"] for r in cat_rows)
    breakdown_html = "".join(
        f"<tr><td>{esc(ch)}</td><td>{cnt}</td><td>{cnt/len(cat_rows)*100:.1f}%</td></tr>"
        for ch,cnt in sorted(cb.items(), key=lambda x:-x[1])
    )
    if has_detail:
        log_html = "".join(
            "<tr><td>" + esc(r['timestamp']) + "</td><td>" + gm_name(r['character_name']) + "</td>"
            "<td>" + esc(char_initials.get(r['character_name'],'')) + "</td>"
            "<td class='cmd'>" + cmd_html(r['command'], r.get('category','')) + "</td>"
            + (
                "<td class='" + detail_class(r['category'],r['detail'],r.get('command','')) + "'"
                + (" style='cursor:pointer;text-decoration:underline dotted' onclick='lookupItem(" + give_item_id(r) + ")'" if give_item_id(r) else "")
                + ">" + esc(r['detail']) + "</td>"
            )
            + "</tr>"
            for r in cat_rows
        )
        log_head = "<tr><th>Timestamp</th><th>Character</th><th>Account</th><th>Command</th><th>Item / Target</th></tr>"
    else:
        log_html = "".join(
            f"<tr><td>{esc(r['timestamp'])}</td><td>{gm_name(r['character_name'])}</td><td>{esc(char_initials.get(r['character_name'],''))}</td><td class='cmd'>{cmd_html(r['command'], r.get('category',''))}</td></tr>"
            for r in cat_rows
        )
        log_head = "<tr><th>Timestamp</th><th>Character</th><th>Account</th><th>Command</th></tr>"
    cat_panes += f"""
<div class="tab-pane" id="cat-{cid}">
  <h3>{badge(cat)} {esc(cat)} <span class="sub">({len(cat_rows)} commands)</span></h3>
  <table class="summary-table"><thead><tr><th>Character</th><th>Count</th><th>%</th></tr></thead><tbody>{breakdown_html}</tbody></table>
  <h4>Command Log</h4>
  <input class="search-bar" type="text" placeholder="Filter..." oninput="filterTable('tbl-cat-{cid}',this.value)">
  <div class="table-wrap"><table id="tbl-cat-{cid}"><thead>{log_head}</thead><tbody>{log_html}</tbody></table></div>
</div>"""

# Char panes
char_panes = ""
for char in chars:
    cr  = [r for r in cmds if r["character_name"]==char]
    cc  = collections.Counter(r["category"] for r in cr)
    breakdown = "".join(
        f"<tr><td>{badge(cat)}</td><td>{cnt}</td><td>{cnt/len(cr)*100:.1f}%</td></tr>"
        for cat,cnt in sorted(cc.items(),key=lambda x:-x[1])
    )
    cmd_rows = "".join(
        f"<tr><td>{esc(r['timestamp'])}</td><td>{badge(r['category'])}</td>"
        f"<td class='cmd'>{cmd_html(r['command'], r.get('category',''))}</td>"
        f"<td>{detail_html(r)}</td></tr>"
        for r in cr
    )
    sid = char.replace(" ","_")
    char_panes += f"""
<div class="tab-pane" id="char-{sid}">
  <h3>{esc(char)} <span class="sub">({len(cr)} commands)</span></h3>
  <table class="summary-table"><thead><tr><th>Type</th><th>Count</th><th>%</th></tr></thead><tbody>{breakdown}</tbody></table>
  <h4>Full Log</h4>
  <input class="search-bar" type="text" placeholder="Filter..." oninput="filterTable('tbl-{sid}',this.value)">
  <div class="table-wrap"><table id="tbl-{sid}">
    <thead><tr><th>Timestamp</th><th>Category</th><th>Command</th><th>Item / Target</th></tr></thead>
    <tbody>{cmd_rows}</tbody>
  </table></div>
</div>"""

summary_chars = "".join(
    f"<tr><td>{gm_name(ch)}</td><td>{esc(char_initials.get(ch,''))}</td><td>{cnt}</td><td>{cnt/total*100:.1f}%</td></tr>"
    for ch,cnt in sorted(by_char.items(),key=lambda x:-x[1])
    if cnt >= 10
)
summary_cats = "".join(
    f"<tr><td>{badge(cat)}</td><td>{cnt}</td><td>{cnt/total*100:.1f}%</td></tr>"
    for cat,cnt in sorted(by_cat.items(),key=lambda x:-x[1])
)
all_rows_html = "".join(
    f"<tr><td>{esc(r['timestamp'])}</td><td>{gm_name(r['character_name'])}</td>"
    f"<td>{esc(char_initials.get(r['character_name'],''))}</td>"
    f"<td>{badge(r['category'])}</td><td class='cmd'>{cmd_html(r['command'], r.get('category',''))}</td>"
    f"<td>{detail_html(r)}</td></tr>"
    for r in cmds
)
# Bots pane HTML
bots_class_summary = "".join(
    f"<tr><td>{esc(cls)}</td><td>{cnt}</td></tr>"
    for cls, cnt in sorted(bots_by_class.items(), key=lambda x: -x[1])
) or '<tr><td colspan="2" style="color:#888;text-align:center">No bots found</td></tr>'

bots_owner_sections = ""
for acct in sorted(bots_by_account.keys()):
    acct_bots = bots_by_account[acct]
    initials  = account_initials(acct, 999)
    rows = "".join(
        f"<tr>"
        f"<td style='color:#888;font-size:.8rem'>{esc(r['owner_name'])}</td>"
        "<td class='item-name' style='cursor:pointer;text-decoration:underline dotted' onclick=\"lookupInventory('" + r['bot_name'].replace("'","&#39;") + "','bot')\">" + esc(r['bot_name']) + "</td>"
        f"<td><span class='badge cat-{esc(r['class_name'].lower().replace(' ','-'))}'>{esc(r['class_name'])}</span></td>"
        f"<td>{esc(r['race_name'])}</td>"
        f"<td>{esc(r['gender_name'])}</td>"
        f"<td style='text-align:center'>{esc(r.get('bot_level',''))}</td>"
        f"<td style='text-align:right'>{esc(r.get('bot_hp',''))}</td>"
        f"<td style='text-align:right'>{esc(r.get('bot_mana',''))}</td>"
        f"</tr>"
        for r in sorted(acct_bots, key=lambda x: (x["owner_name"], x["bot_name"]))
    )
    bots_owner_sections += f"""
  <h4 style="margin:1.25rem 0 .5rem">{esc(initials)} <span style="font-size:.8rem;color:#888;font-weight:400">({esc(acct)})</span> <span style="font-weight:400;color:#888;font-size:.85rem">— {len(acct_bots)} bots</span></h4>
  <div class="table-wrap" style="margin-bottom:.5rem"><table>
    <thead><tr><th>Character</th><th>Bot Name</th><th>Class</th><th>Race</th><th>Gender</th><th>Level</th><th>HP</th><th>Mana</th></tr></thead>
    <tbody>{rows}</tbody>
  </table></div>"""

kill_log_html = "".join(
    f"<tr><td>{esc(r['timestamp'])}</td><td>{gm_name(r['character_name'])}</td>"
    f"<td class='{kill_target_class(r['target_name'])}'>{esc(r['target_name'])}</td>"
    f"<td>{esc(r['target_id'])}</td></tr>"
    for r in kills
)
kill_summary_html = "".join(
    f"<tr><td class='{kill_target_class(t)}'>{esc(t)}</td><td>{cnt}</td></tr>"
    for t,cnt in kill_by_target.most_common()
)
item_log_html = "".join(
    f"<tr><td>{esc(r['timestamp'])}</td><td>{gm_name(r['character_name'])}</td>"
    f"<td>{esc(r['item_id'])}</td>"
    "<td class='" + item_class(r['rarity']) + "' style='cursor:pointer;text-decoration:underline dotted' onclick='lookupItem(" + r['item_id'] + ")'>" + esc(r['item_name']) + "</td>"
    f"<td>{rarity_badge(r['rarity'])}</td>"
    f"<td style='text-align:center'>{esc(str(r['quantity']))}</td>"
    f"<td class='cmd'>{esc(r['raw_command'])}</td></tr>"
    for r in items
)
item_summary_html = "".join(
    "<tr><td class='" + item_class(item_rarity_lookup.get(name_to_id.get(nm,""),"common")) + "' style='cursor:pointer;text-decoration:underline dotted' onclick='lookupItem(" + name_to_id.get(nm,"0") + ")'>" + esc(nm) + "</td><td>" + str(cnt) + "</td></tr>"
    for nm,cnt in item_by_name.most_common()
)

legendary_count = sum(1 for r in items if r["rarity"] in ("orange","purple"))
notable_count   = sum(1 for r in items if r["rarity"] in ("blue","green"))

# Build characters pane HTML
def char_row(r):
    nm   = r["name"].replace("'", "\'")
    clk  = f"lookupInventory('{nm}','character')"
    if r["is_gm"]:
        name_html = f'<span style="color:#fb923c;font-weight:700;cursor:pointer" onclick="{clk}">{esc(r["name"])}</span>'
    else:
        name_html = f'<span style="cursor:pointer;text-decoration:underline dotted" onclick="{clk}">{esc(r["name"])}</span>'
    guild = f'<span style="color:var(--text3);font-size:.78rem">{esc(r["guild_name"])}</span>' if r.get("guild_name") else ''
    gm_badge = '<span class="badge cat-gm-toggle" style="font-size:.68rem;padding:.1rem .35rem;vertical-align:middle">GM</span>' if r["is_gm"] else ''
    return (
        f"<tr>"
        f"<td>{name_html} {gm_badge}</td>"
        f"<td style='color:var(--text2)'>{esc(r['class_name'])}</td>"
        f"<td style='text-align:center'>{esc(r.get('level',''))}</td>"
        f"<td style='color:var(--text3);font-size:.8rem'>{esc(r['race_name'])}</td>"
        f"<td style='color:var(--text2);font-size:.8rem'>{esc(r.get('account_name',''))}</td>"
        f"<td style='color:var(--text3);font-size:.78rem'>{guild}</td>"
        f"<td style='color:var(--text3);font-size:.78rem'>{esc(r.get('last_login',''))}</td>"
        f"</tr>"
    )

import json as _json
cmds_json    = _json.dumps([{"timestamp":r.get("timestamp",""),"character_name":r.get("character_name",""),"category":r.get("category","")} for r in cmds])

chars_html = "".join(char_row(r) for r in chars_raw) or '<tr><td colspan="7" style="color:#888;text-align:center">No character data</td></tr>'
# ---- Deaths HTML ----
def death_row(r):
    ts   = r.get("timestamp","")
    char = r.get("character_name","")
    lvl  = r.get("level","?")
    kb   = r.get("killed_by","") or ""
    zone = r.get("zone","") or ""
    # Environmental deaths (falls, drowning, etc.) have no killer — show dash
    kb_display   = (f'<td style="color:#f87171;font-weight:600">{esc(kb)}</td>'
                    if kb and kb.lower() not in ("unknown","null","")
                    else '<td style="color:var(--text3)">—</td>')
    zone_display = (f'<td style="color:var(--text3)">{esc(zone)}</td>'
                    if zone and zone.lower() not in ("unknown","null","")
                    else '<td style="color:var(--text3)">—</td>')
    return (f'<tr><td>{ts}</td><td>{char}</td><td style="text-align:center">{lvl}</td>'
            f'{kb_display}{zone_display}</tr>')

deaths_html = "".join(death_row(r) for r in deaths_raw) or '<tr><td colspan="5" style="color:#888;text-align:center">No death records</td></tr>'


# Pre-build online/offline HTML blocks
online_html = "".join(
    '<div style="display:flex;align-items:center;gap:.5rem;background:var(--bg3);border:1px solid #22d3a055;border-radius:6px;padding:.4rem .8rem;margin:.25rem .25rem .25rem 0;font-size:.85rem;min-width:0;max-width:100%;overflow:hidden">'
    + '<span style="width:8px;height:8px;border-radius:50%;background:#22d3a0;box-shadow:0 0 5px #22d3a0;flex-shrink:0"></span>'
    + (f'<strong style="color:#fb923c">{esc(r["character_name"])}</strong>' if r["character_name"] in gm_names else f'<strong style="color:var(--accent)">{esc(r["character_name"])}</strong>')
    + f'<span style="color:var(--text2)"> {esc(r.get("class_name","?"))} L{esc(r.get("level",""))}</span>'
    + f'<span style="color:var(--text);font-size:.8rem"> &#8212; {esc(r["zone_name"])}</span>'
    + f'<span style="color:#22d3a0;font-size:.75rem;margin-left:.35rem">({esc(format_duration(r.get("seconds_online","0")))})</span>'
    + '</div>'
    for r in online
) or '<span style="color:var(--text3);font-size:.85rem">No players currently online</span>'

offline_html = "".join(
    '<div style="background:var(--bg3);border:1px solid var(--border);border-radius:6px;padding:.5rem .8rem;margin:.25rem 0;font-size:.85rem;display:flex;align-items:center;gap:.5rem;flex-wrap:wrap;min-width:0">'
    + '<span style="width:8px;height:8px;border-radius:50%;background:#4a5070;flex-shrink:0"></span>'
    + (f'<strong style="color:#fb923c">{esc(r["character_name"])}</strong>' if r.get("is_gm","0") == "1" else f'<strong style="color:var(--text)">{esc(r["character_name"])}</strong>')
    + f'<span style="color:var(--text2)"> {esc(r.get("class_name","?"))} L{esc(r.get("level",""))}</span>'
    + f'<span style="color:var(--text3);font-size:.78rem;margin-left:auto">offline {esc(format_duration(r.get("seconds_offline","0")))} &mdash; {esc(r.get("last_login_time","?"))}</span>'
    + '</div>'
    for r in offline
) or '<span style="color:var(--text3);font-size:.85rem">No recent players</span>'

# Embed commands as JS for client-side new-since-last-view calculation
cmds_js = json.dumps([
    [r["timestamp"], r["character_name"],
     char_initials.get(r["character_name"],""),
     r["category"], r["command"], r["detail"],
     detail_class(r["category"], r["detail"], r.get("command",""))]
    for r in cmds
])
legendary_given = [r for r in impact if r["rarity"] in ("orange","purple")]
notable_given   = [r for r in impact if r["rarity"] in ("blue","green")]

# Web new-commands display is handled entirely by client-side JS
new_cmds_card_extra = 'id="new-card" style="border:' + delta_border + ';cursor:pointer" onclick="showTab(\'new-cmds\',null)"'
new_cmds_nav_btn    = '<button class="tab-btn" id="new-nav-btn" style="display:none;border-color:#F59E0B;color:#D97706" onclick="showTab(\'new-cmds\',this)">&#128276; New (<span id="new-nav-count">0</span>)</button>'

new_cmds_pane = (
    '<div class="tab-pane" id="new-cmds">'
    '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1rem">'
    '<h3 style="margin:0">&#128276; New Commands <span class="sub" id="new-cmds-sub">since last clear</span></h3>'
    '<button onclick="clearNewCmds()" style="background:var(--accent);color:#fff;border:none;border-radius:6px;padding:.4rem 1rem;cursor:pointer;font-size:.82rem;font-weight:600">&#10003; Mark all seen</button>'
    '</div>'
    '<input class="search-bar" type="text" placeholder="Filter..." oninput="filterTable(\'tbl-new-cmds\',this.value)">'
    '<div class="table-wrap"><table id="tbl-new-cmds">'
    '<thead><tr><th>Timestamp</th><th>Character</th><th>Account</th><th>Category</th><th>Command</th><th>Item / Target</th></tr></thead>'
    '<tbody id="new-cmds-body"><tr><td colspan="6" style="text-align:center;color:var(--text3)">Loading...</td></tr></tbody>'
    '</table></div></div>'
)


# ── Summary tab highlights ────────────────────────────────────────────────────
hl_kills = "".join(
    f"<tr><td class='{kill_target_class(t)}' style='white-space:nowrap;overflow:hidden;"
    f"text-overflow:ellipsis;max-width:180px'>{esc(t)}</td>"
    f"<td style='text-align:right;font-weight:700;color:var(--text);padding-left:.5rem'>{cnt}</td></tr>"
    for t, cnt in kill_by_target.most_common(5)
) or '<tr><td colspan="2" style="color:var(--text3);text-align:center;font-size:.82rem;padding:.5rem">No kill data</td></tr>'

hl_items = "".join(
    "<tr><td class='" + item_class(item_rarity_lookup.get(name_to_id.get(nm,""),"grey")) +
    "' style='cursor:pointer;text-decoration:underline dotted;white-space:nowrap;"
    "overflow:hidden;text-overflow:ellipsis;max-width:180px'"
    " onclick='lookupItem(" + (name_to_id.get(nm,"0") or "0") + ")'>" + esc(nm) +
    "</td><td style='text-align:right;font-weight:700;color:var(--text);padding-left:.5rem'>" +
    str(cnt) + "</td></tr>"
    for nm, cnt in item_by_name.most_common(5)
) or '<tr><td colspan="2" style="color:var(--text3);text-align:center;font-size:.82rem;padding:.5rem">No item data</td></tr>'

hl_recip = "".join(
    f"<tr><td style='cursor:pointer;text-decoration:underline dotted'"
    f" onclick=\"lookupInventory('{esc(ch)}','character')\"><span style='color:var(--accent)'>{esc(ch)}</span></td>"
    f"<td style='text-align:right;font-weight:700;color:var(--text);padding-left:.5rem'>{cnt}</td></tr>"
    for ch, cnt in impact_by_recipient.most_common(5)
) or '<tr><td colspan="2" style="color:var(--text3);text-align:center;font-size:.82rem;padding:.5rem">No recipient data</td></tr>'

_pk_day_val, _pk_day_cnt = (max(daily_counts.items(), key=lambda x: x[1]) if daily_counts else ("N/A", 0))
_pk_h, _pk_h_cnt         = (peak_hours[0] if peak_hours and peak_hours[0][1] > 0 else (0, 0))
_pk_ampm                 = f"{_pk_h % 12 or 12}{'am' if _pk_h < 12 else 'pm'}"
_pk_period_val, _pk_period_cnt = (max(bucket_counts.items(), key=lambda x: x[1]) if bucket_counts else ("N/A", 0))
_max_cat = max(by_cat.values()) if by_cat else 1

hl_cats = "".join(
    f"<tr><td style='padding:.25rem .5rem'>{badge(cat)}</td>"
    f"<td style='text-align:right;font-weight:700;color:var(--text);padding:.25rem .5rem;"
    f"font-size:.82rem'>{cnt}</td>"
    f"<td style='padding:.25rem .5rem'><div style='height:6px;border-radius:3px;"
    f"background:var(--accent);width:{int(cnt/_max_cat*100)}%'></div></td></tr>"
    for cat, cnt in by_cat.most_common(5)
)

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>GM Command Usage</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{{
  --bg:#0d0f18;--bg2:#131620;--bg3:#0f111a;--bg4:#070910;
  --border:#1e2235;--border2:#252840;
  --text:#dde1ee;--text2:#9aa4c8;--text3:#6070a0;
  --accent:#22d3a0;--accent-bg:#0f2e28;--accent-border:#1a4a40;
  --card-bg:#131620;--header-bg:#070910;
  --tab-bg:#131620;--tab-hover:#1a1e2e;
  --td-alt:#0f111a;--td-hover:#1a1e2e;
  --input-bg:#0d0f18;--chart-grid:#1e2235;
  --cmd-color:#6a8090;
}}
[data-theme="light"]{{
  --bg:#f0f2f5;--bg2:#ffffff;--bg3:#f9f9f9;--bg4:#2d2d2d;
  --border:#e0e0e0;--border2:#d0d0d0;
  --text:#2d2d2d;--text2:#555;--text3:#888;
  --accent:#1a7b72;--accent-bg:#e8f5f2;--accent-border:#1a7b72;
  --card-bg:#ffffff;--header-bg:#2d2d2d;
  --tab-bg:#ffffff;--tab-hover:#e8f5f2;
  --td-alt:#f9f9f9;--td-hover:#e8f5f2;
  --input-bg:#ffffff;--chart-grid:#e5e7eb;
  --cmd-color:#555;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;background:var(--bg);color:var(--text);font-size:14px;transition:background .2s,color .2s}}
header{{background:var(--header-bg);color:#fff;padding:.9rem 2rem;display:flex;align-items:center;justify-content:space-between;gap:1rem;border-bottom:1px solid var(--border);flex-wrap:wrap}}
header h1{{font-size:1.1rem;font-weight:700;white-space:nowrap;color:#e0e4f0}}
header .meta{{font-size:.8rem;color:var(--text3);text-align:right}}
.theme-btn{{background:transparent;border:1px solid var(--border2);border-radius:20px;padding:.3rem .8rem;cursor:pointer;font-size:.8rem;color:#aaa;white-space:nowrap;transition:all .2s}}
.theme-btn:hover{{border-color:var(--accent);color:var(--accent)}}
.container{{max-width:1400px;margin:0 auto;padding:1.5rem;overflow-x:hidden}}
.stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:1rem;margin-bottom:1.5rem}}
.stat-card{{background:var(--card-bg);border-radius:8px;padding:1rem 1.25rem;border:.5px solid var(--border);transition:background .2s}}
.stat-card .val{{font-size:1.6rem;font-weight:700;color:var(--accent);line-height:1.1}}
.stat-card .lbl{{font-size:.7rem;color:var(--text3);margin-top:.25rem;text-transform:uppercase;letter-spacing:.04em}}
.stat-card .sub{{font-size:.65rem;color:var(--text3);margin-top:.1rem}}
.nav-card{{background:var(--card-bg);border:.5px solid var(--border);border-radius:10px;padding:.9rem 1.25rem;margin-bottom:1rem;display:flex;flex-direction:column;gap:.75rem}}
.nav-row{{display:flex;align-items:flex-start;gap:1.5rem;flex-wrap:wrap}}
.nav-section{{display:flex;flex-direction:column;gap:.35rem;min-width:0}}
.nav-section.fill{{flex:1;min-width:0;overflow:hidden}}
.nav-divider{{width:1px;background:var(--border);align-self:stretch;flex-shrink:0;margin:0 .25rem}}
.nav-label{{font-size:.62rem;text-transform:uppercase;letter-spacing:.1em;color:var(--text3);font-weight:600}}
.tabs{{display:flex;gap:.3rem;flex-wrap:wrap}}
.tab-btn{{background:transparent;border:.5px solid var(--border);border-radius:5px;padding:.3rem .7rem;cursor:pointer;font-size:.78rem;color:var(--text2);transition:all .15s;display:inline-flex;align-items:center;gap:.3rem;white-space:nowrap;line-height:1.4}}
.tab-btn:hover{{background:var(--tab-hover);border-color:var(--accent);color:var(--text)}}
.tab-btn.active{{background:var(--accent);color:#fff;border-color:var(--accent)}}
.tab-btn.active .badge{{background:rgba(255,255,255,.2)!important;color:#fff!important;border-color:transparent!important}}
.tab-btn.filter-btn.active{{background:var(--bg3);color:var(--text);border-color:var(--accent);box-shadow:0 0 0 1px var(--accent)}}
@media(max-width:800px){{
  .nav-row{{flex-direction:column}}
  .nav-divider{{display:none}}
  .two-col{{grid-template-columns:1fr}}
  .three-col{{grid-template-columns:1fr 1fr}}
  .stats{{grid-template-columns:repeat(auto-fit,minmax(110px,1fr))}}
  .container{{padding:1rem .75rem}}
  header{{padding:.75rem 1rem}}
  header h1{{font-size:.95rem}}
  .server-bar{{padding:.4rem 1rem;gap:.75rem}}
  .tab-pane{{padding:1rem .75rem}}
  .chart-wrap{{padding:1rem .75rem}}
  h3{{font-size:1rem}}
}}
@media(max-width:480px){{
  .two-col{{grid-template-columns:1fr}}
  .three-col{{grid-template-columns:1fr}}
  .stats{{grid-template-columns:repeat(2,1fr)}}
  .stat-card .val{{font-size:1.3rem}}
  .tab-btn{{font-size:.72rem;padding:.25rem .5rem}}
  table{{font-size:.76rem}}
  th,td{{padding:.35rem .5rem}}
}}
.tab-pane{{display:none;background:var(--card-bg);border-radius:8px;padding:1.5rem;border:.5px solid var(--border);margin-top:.75rem;transition:background .2s}}
.tab-pane.active{{display:block}}
.search-bar{{width:100%;padding:.45rem .75rem;border:.5px solid var(--border);border-radius:6px;font-size:.88rem;margin-bottom:.75rem;outline:none;background:var(--input-bg);color:var(--text)}}
.search-bar:focus{{border-color:var(--accent);box-shadow:0 0 0 2px #1a7b7233}}
.search-bar::placeholder{{color:var(--text3)}}
table{{width:100%;border-collapse:collapse;font-size:.83rem}}
th{{background:var(--bg4);color:var(--text2);padding:.55rem .75rem;text-align:left;white-space:nowrap;position:sticky;top:0;z-index:1;border-bottom:1px solid var(--border)}}
td{{padding:.45rem .75rem;border-bottom:1px solid var(--border);vertical-align:top;color:var(--text);word-break:break-word;max-width:300px}}
tr:nth-child(even) td{{background:var(--td-alt)}}
tr:hover td{{background:var(--td-hover)}}
.cmd{{font-family:monospace;font-size:.78rem;word-break:break-all;color:var(--cmd-color);white-space:pre-wrap}}
.spell-link{{color:#a78bfa;cursor:pointer;text-decoration:underline dotted;font-weight:600}}
.spell-link:hover{{color:#c4b5fd}}
.item-link{{color:var(--accent);cursor:pointer;text-decoration:underline dotted;font-weight:600}}
.item-link:hover{{color:#fff}}
.target{{font-weight:600;color:var(--accent)}}
.named-target{{color:#c084fc!important;font-weight:700}}
.item-name{{font-weight:600;color:var(--accent)}}
.item-orange{{font-weight:700;color:#ff8000;text-shadow:0 0 8px #ff800066}}
.item-purple{{font-weight:700;color:#c084fc}}
.item-blue{{font-weight:600;color:#60a5fa}}
.item-green{{font-weight:600;color:#4ade80}}
.item-white{{font-weight:500;color:#94a3b8}}
.item-grey{{font-weight:400;color:var(--text3)}}
.summary-table{{width:100%;margin-bottom:1.5rem;min-width:0}}
.table-wrap{{max-height:500px;overflow-y:auto;overflow-x:auto;border-radius:6px;border:.5px solid var(--border)}}
.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:1.5rem;margin-bottom:2rem}}
.three-col{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:1.5rem;margin-bottom:2rem}}
.section-head{{background:var(--accent-bg);color:var(--accent);padding:.5rem .75rem;border-radius:6px 6px 0 0;font-weight:700;font-size:.88rem;border:1px solid var(--accent-border);border-bottom:none;width:100%}}
.chart-wrap{{background:var(--card-bg);border-radius:8px;border:.5px solid var(--border);padding:1.25rem;margin-bottom:1.5rem;transition:background .2s}}
.chart-wrap h4{{margin-bottom:1rem;font-size:.9rem;color:var(--text2)}}
.badge{{display:inline-block;padding:.15rem .45rem;border-radius:4px;font-size:.72rem;font-weight:600;white-space:nowrap}}
.cat-give-item{{background:#0f2e1a;color:#4ade80;border:1px solid #1a4a2a}}
.cat-find-item{{background:#0f1e35;color:#60a5fa;border:1px solid #1a3050}}
.cat-kill{{background:#2e0f0f;color:#f87171;border:1px solid #4a1a1a}}
.cat-buff,.cat-heal,.cat-mana{{background:#2a200a;color:#fbbf24;border:1px solid #4a380a}}
.cat-set-level,.cat-set-aa{{background:#1e102e;color:#c084fc;border:1px solid #3a1a50}}
.cat-zone,.cat-goto,.cat-summon{{background:#0a1e28;color:#38bdf8;border:1px solid #0a3040}}
.cat-gm-toggle{{background:#2a1a08;color:#fb923c;border:1px solid #4a2a08}}
.cat-spawn-npc,.cat-npc-edit{{background:#1e0a2e;color:#e879f9;border:1px solid #3a0a50}}
.cat-other,.cat-unknown{{background:var(--bg3);color:var(--text3);border:1px solid var(--border)}}
.rarity-orange{{background:#ff800022;color:#ff8000;border:1px solid #ff800055}}
.rarity-purple{{background:#c084fc22;color:#c084fc;border:1px solid #c084fc55}}
.rarity-blue{{background:#60a5fa22;color:#60a5fa;border:1px solid #60a5fa55}}
.rarity-green{{background:#4ade8022;color:#4ade80;border:1px solid #4ade8055}}
.rarity-white{{background:#94a3b822;color:#94a3b8;border:1px solid #94a3b855}}
.rarity-grey{{background:var(--border)22;color:var(--text3);border:1px solid var(--border)}}
.server-bar{{background:var(--bg4);border-bottom:1px solid var(--border);padding:.4rem 2rem;display:flex;align-items:center;gap:1.5rem;font-size:.78rem;color:var(--text3);flex-wrap:wrap}}
.server-bar .sb-item{{display:flex;align-items:center;gap:.35rem;flex-shrink:0}}
.server-bar .sb-val{{color:var(--text2);font-weight:600}}
.proc-dot{{width:7px;height:7px;border-radius:50%;display:inline-block;flex-shrink:0}}
.proc-ok{{background:#22d3a0;box-shadow:0 0 4px #22d3a0}}
.proc-warn{{background:#fbbf24}}
.proc-down{{background:#f87171}}
h3 .sub{{font-size:.85rem;font-weight:400;color:var(--text3)}}
h4{{font-size:.9rem;color:var(--text2);margin:.75rem 0 .4rem}}
.item-modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:2000;align-items:center;justify-content:center;padding:1rem}}
.item-modal-overlay.open{{display:flex!important}}
.item-card{{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:1.5rem 1.75rem;width:100%;max-width:460px;max-height:80vh;overflow-y:auto;position:relative;box-shadow:0 20px 60px rgba(0,0,0,.5)}}
@media(max-width:480px){{
  .item-card{{padding:1rem;max-height:90vh}}
  .item-modal-overlay{{padding:.5rem}}
}}
.item-flag{{font-size:.68rem;font-weight:700;padding:.12rem .4rem;border-radius:3px;display:inline-block;letter-spacing:.04em}}
.health-bar{{width:50px;height:5px;background:var(--bg3);border-radius:3px;overflow:hidden;display:inline-block;vertical-align:middle;margin-left:.25rem}}
.health-fill{{height:100%;border-radius:3px;transition:width .4s}}
.zone-chip{{display:inline-flex;align-items:center;gap:.4rem;background:var(--bg3);border:1px solid var(--border);border-radius:5px;padding:.25rem .65rem;margin:.2rem;font-size:.8rem}}
.inv-slot{{display:grid;grid-template-columns:90px 1fr;gap:.25rem .6rem;align-items:center;padding:.3rem .4rem;border-radius:5px;font-size:.82rem}}
.inv-slot:hover{{background:var(--bg3)}}
.inv-slot-label{{color:var(--text3);font-size:.72rem;text-transform:uppercase;letter-spacing:.05em}}
.inv-item-name{{cursor:pointer;text-decoration:underline dotted}}
.inv-empty{{color:var(--text3);font-style:italic}}
.inv-mini-stats{{font-size:.72rem;color:var(--text3);display:flex;gap:.4rem;flex-wrap:wrap;margin-top:.1rem}}
</style>
</head>
<body>
<header>
  <h1>&#9876; GM Command Usage</h1>
  <div style="display:flex;align-items:center;gap:1rem">
    <span class="meta">Generated: {generated}<br>Total: {total}</span>
    <button class="theme-btn" id="theme-toggle" onclick="toggleTheme()">&#9728; Light</button>
  </div>
</header>
<div class="server-bar" id="server-bar">
  <span class="sb-item">&#127758; <span class="sb-val" id="sb-server">—</span></span>
  <span class="sb-item">&#8593; <span class="sb-val" id="sb-uptime">—</span></span>
  <span class="sb-item" id="sb-world-wrap"><span class="proc-dot proc-down" id="sb-world-dot"></span>World</span>
  <span class="sb-item" id="sb-zone-wrap"><span class="proc-dot proc-down" id="sb-zone-dot"></span>Zones: <span class="sb-val" id="sb-zones">—</span></span>
  <span class="sb-item" id="sb-ucs-wrap"><span class="proc-dot proc-down" id="sb-ucs-dot"></span>UCS</span>
  <span class="sb-item">CPU: <span class="sb-val" id="sb-cpu">—</span><span class="health-bar"><span class="health-fill" id="sb-cpu-bar" style="background:#22d3a0;width:0%"></span></span></span>
  <span class="sb-item">RAM: <span class="sb-val" id="sb-ram">—</span><span class="health-bar"><span class="health-fill" id="sb-ram-bar" style="background:#60a5fa;width:0%"></span></span></span>
  <span class="sb-item">&#128100; <span class="sb-val" id="sb-players">—</span> online</span>
  <span class="sb-item" style="margin-left:auto;font-size:.7rem" id="sb-updated"></span>
</div>
<div class="container">
  <div class="stats">
    <div class="stat-card"><div class="val">{total}</div><div class="lbl">Total Commands</div></div>
    <div class="stat-card" {new_cmds_card_extra}>
      <div class="val" id="new-count-val" style="color:#F59E0B">—</div>
      <div class="lbl">New Commands</div>
      <div class="sub" id="new-count-sub">since last clear</div>
    </div>
    <div class="stat-card"><div class="val">{len(kills)}</div><div class="lbl">Kill Events</div></div>
    <div class="stat-card"><div class="val">{len(items)}</div><div class="lbl">Items Given</div></div>
    <div class="stat-card"><div class="val">{len(impact_by_recipient)}</div><div class="lbl">Players Impacted</div></div>
    <div class="stat-card"><div class="val" id="sv-accounts">—</div><div class="lbl">Accounts</div></div>
    <div class="stat-card" style="cursor:pointer" onclick="showTab('characters',document.querySelector('.tab-btn[onclick*=characters]'))"><div class="val" id="sv-characters">—</div><div class="lbl">Characters</div></div>
    <div class="stat-card"><div class="val" id="sv-guilds">—</div><div class="lbl">Guilds</div></div>
  </div>

  <div class="nav-card">
    <div class="nav-row">
      <div class="nav-section fill">
        <div class="nav-label">&#128203; Overview</div>
        <div class="tabs">
          <button class="tab-btn active" onclick="showTab('summary',this)">Summary</button>
          {new_cmds_nav_btn}
          <button class="tab-btn" onclick="showTab('trends',this)">&#128200; Trends</button>
          <button class="tab-btn" onclick="showTab('impact',this)">&#128100; Impact</button>
          <button class="tab-btn" onclick="showTab('bots',this)">&#129302; Bots</button>
          <button class="tab-btn" onclick="showTab('details',this)">&#128481; Kills &amp; Items</button>
          <button class="tab-btn" onclick="showTab('all',this)">All Commands</button>
          <button class="tab-btn" onclick="showTab('characters',this);loadLevelChart()">&#128100; Characters</button>
          <button class="tab-btn" onclick="showTab('spire-log',this);if(!window._spireLogLoaded)loadSpireLog('')">&#128220; Spire Log</button>
        </div>
      </div>
      <div class="nav-divider"></div>
      <div class="nav-section">
        <div class="nav-label">&#128100; GM Character</div>
        <div class="tabs">{char_tab_btns}</div>
      </div>
    </div>
    <div class="nav-row">
      <div class="nav-section fill">
        <div class="nav-label">&#127991; Command Type</div>
        <div class="tabs">{cat_tab_btns}</div>
      </div>
    </div>
  </div>

  {new_cmds_pane}

  <div class="tab-pane active" id="summary">
    <h3>&#128994; Online Now <span id="online-pane-count" style="font-size:.75rem;color:var(--text3);font-weight:400"></span></h3>
    <div id="online-pane" style="margin-bottom:.75rem">{online_html}</div>
    <div id="zone-breakdown" style="display:none;margin-bottom:1.5rem">
      <div style="font-size:.65rem;text-transform:uppercase;letter-spacing:.1em;color:var(--text3);margin-bottom:.4rem">&#128205; Who's Where</div>
      <div id="zone-chips"></div>
    </div>
    <div class="two-col" style="align-items:start">
      <div>
        <h3>Commands by Character</h3>
        <table class="summary-table"><thead><tr><th>Character</th><th>Account</th><th>Count</th><th>%</th></tr></thead><tbody>{summary_chars}</tbody></table>
        <h3>Commands by Type</h3>
        <table class="summary-table"><thead><tr><th>Type</th><th>Count</th><th>%</th></tr></thead><tbody>{summary_cats}</tbody></table>
      </div>
      <div>
        <h3>&#128308; Recently Offline</h3>
        {offline_html}
      </div>
    </div>
    <div style="margin-top:2rem">
      <h3>&#128202; Highlights</h3>
      <div class="three-col" style="margin-top:.75rem">
        <div>
          <div class="section-head">&#9876; Top Kills</div>
          <div class="table-wrap" style="max-height:220px"><table>
            <thead><tr><th>Target</th><th style="text-align:right">Count</th></tr></thead>
            <tbody>{hl_kills}</tbody>
          </table></div>
        </div>
        <div>
          <div class="section-head">&#127873; Top Items Given</div>
          <div class="table-wrap" style="max-height:220px"><table>
            <thead><tr><th>Item</th><th style="text-align:right">Count</th></tr></thead>
            <tbody>{hl_items}</tbody>
          </table></div>
        </div>
        <div>
          <div class="section-head">&#128100; Top Recipients</div>
          <div class="table-wrap" style="max-height:220px"><table>
            <thead><tr><th>Character</th><th style="text-align:right">Items</th></tr></thead>
            <tbody>{hl_recip}</tbody>
          </table></div>
        </div>
      </div>
      <div class="two-col" style="margin-top:1rem">
        <div>
          <div class="section-head">&#128336; Peak Activity</div>
          <table style="font-size:.83rem;width:100%"><tbody>
            <tr><td style="color:var(--text3);padding:.3rem .5rem">Most active day</td>
                <td style="font-weight:600;padding:.3rem .5rem">{_pk_day_val}</td>
                <td style="color:var(--accent);padding:.3rem .5rem;text-align:right">{_pk_day_cnt} cmds</td></tr>
            <tr><td style="color:var(--text3);padding:.3rem .5rem">Peak hour</td>
                <td style="font-weight:600;padding:.3rem .5rem">{_pk_h:02d}:00 ({_pk_ampm})</td>
                <td style="color:var(--accent);padding:.3rem .5rem;text-align:right">{_pk_h_cnt} cmds</td></tr>
            <tr><td style="color:var(--text3);padding:.3rem .5rem">Peak period</td>
                <td style="font-weight:600;padding:.3rem .5rem">{_pk_period_val}</td>
                <td style="color:var(--accent);padding:.3rem .5rem;text-align:right">{_pk_period_cnt} cmds</td></tr>
          </tbody></table>
        </div>
        <div>
          <div class="section-head">&#127991; Top Command Types</div>
          <table style="width:100%"><tbody>{hl_cats}</tbody></table>
        </div>
      </div>
    </div>
  </div>

  <div class="tab-pane" id="trends">
    <div class="chart-wrap">
      <h4>Daily Command Volume</h4>
      <canvas id="chart-daily" height="80"></canvas>
    </div>
    <div class="two-col">
      <div class="chart-wrap">
        <h4>Commands by Hour of Day</h4>
        <canvas id="chart-hourly"></canvas>
      </div>
      <div class="chart-wrap">
        <h4>Commands by Type</h4>
        <canvas id="chart-types"></canvas>
      </div>
    </div>
    <div class="three-col">
      <div>
        <div class="section-head">&#128336; Peak Hours Ranked</div>
        <div class="table-wrap"><table>
          <thead><tr><th>Hour</th><th></th><th>Count</th><th>Volume</th></tr></thead>
          <tbody>{peak_hours_html or '<tr><td colspan="4" style="color:#888;text-align:center">No data</td></tr>'}</tbody>
        </table></div>
      </div>
      <div>
        <div class="section-head">&#127775; Time of Day</div>
        <div class="table-wrap"><table>
          <thead><tr><th>Period</th><th>Count</th><th>Volume</th></tr></thead>
          <tbody>{bucket_html}</tbody>
        </table></div>
      </div>
      <div>
        <div class="section-head">&#128197; Day of Week</div>
        <div class="table-wrap"><table>
          <thead><tr><th>Day</th><th>Count</th><th>Volume</th></tr></thead>
          <tbody>{dow_html}</tbody>
        </table></div>
      </div>
    </div>

    <h3 style="margin-top:1.5rem">&#128197; Player Activity Heatmap
      <span style="font-size:.8rem;font-weight:400;color:var(--text3);margin-left:.5rem">Commands by day &amp; hour</span>
    </h3>
    <div id="activity-heatmap" style="background:var(--bg2);border:1px solid var(--border);border-radius:6px;padding:1rem">
      <span style="color:var(--text3)">Loading...</span>
    </div>

    <h3 style="margin-top:1.5rem">&#9654; Play Sessions
      <span style="font-size:.8rem;font-weight:400;color:var(--text3);margin-left:.5rem">Groups of activity separated by 30+ min gaps</span>
    </h3>
    <div id="session-table">
      <span style="color:var(--text3)">Loading...</span>
    </div>
  </div>

  <div class="tab-pane" id="impact">
    <div class="two-col">
      <div>
        <div class="section-head">&#128100; Items Received by Character</div>
        <div class="table-wrap"><table>
          <thead><tr><th>Character</th><th>Items Received</th></tr></thead>
          <tbody>{impact_by_recipient_html}</tbody>
        </table></div>
      </div>
      <div>
        <div class="section-head">&#127873; Legendary &amp; Notable Items Given</div>
        <div class="table-wrap"><table>
          <thead><tr><th>Recipient</th><th>Item</th><th>Rarity</th></tr></thead>
          <tbody>{"".join("<tr><td class='item-name'>" + esc(r['recipient']) + "</td><td class='" + item_class(r['rarity']) + "' style='cursor:pointer;text-decoration:underline dotted' onclick='lookupItem(" + r['item_id'] + ")'>" + esc(r['item_name']) + "</td><td>" + rarity_badge(r['rarity']) + "</td></tr>" for r in sorted(legendary_given+notable_given, key=lambda x: (["orange","purple","blue","green"].index(x['rarity']) if x['rarity'] in ["orange","purple","blue","green"] else 9, x['recipient']))) or '<tr><td colspan="3" style="color:#888;text-align:center">No notable items found</td></tr>'}</tbody>
        </table></div>
      </div>
    </div>
    <h4>Full Impact Log</h4>
    <input class="search-bar" type="text" placeholder="Filter by character, item, GM..." oninput="filterTable('tbl-impact',this.value)">
    <div class="table-wrap"><table id="tbl-impact">
      <thead><tr><th>Timestamp</th><th>GM Character</th><th>Recipient</th><th>Item ID</th><th>Item Name</th><th>Rarity</th></tr></thead>
      <tbody>{impact_log_html}</tbody>
    </table></div>
  </div>

  <div class="tab-pane" id="bots">
    <div class="two-col" style="margin-bottom:1.5rem">
      <div>
        <div class="section-head">&#129302; Bots by Account</div>
        <div class="table-wrap"><table>
          <thead><tr><th>Class</th><th>Count</th></tr></thead>
          <tbody>{bots_class_summary}</tbody>
        </table></div>
      </div>
      <div class="stat-card" style="display:flex;flex-direction:column;justify-content:center;align-items:center;gap:.5rem">
        <div style="font-size:3rem;font-weight:700;color:#1a7b72">{total_bots}</div>
        <div style="font-size:.8rem;color:#888;text-transform:uppercase;letter-spacing:.05em">Total Bots</div>
        <div style="font-size:.8rem;color:#888">{len(bots_by_account)} account(s)</div>
      </div>
    </div>
    {bots_owner_sections if bots_owner_sections else '<p style="color:#888;text-align:center">No bot data found</p>'}
  </div>

  <div class="tab-pane" id="details">
    <div class="two-col">
      <div><div class="section-head">&#9876; Kill Summary</div>
        <div class="table-wrap"><table><thead><tr><th>Target Name</th><th>Times Killed</th></tr></thead>
        <tbody>{kill_summary_html or '<tr><td colspan="2" style="color:#888;text-align:center">No kill target data</td></tr>'}</tbody></table></div>
      </div>
      <div><div class="section-head">&#127873; Item Summary</div>
        <div class="table-wrap"><table><thead><tr><th>Item Name</th><th>Times Given</th></tr></thead>
        <tbody>{item_summary_html or '<tr><td colspan="2" style="color:#888;text-align:center">No item data</td></tr>'}</tbody></table></div>
      </div>
    </div>
    <h4>Full Kill Log</h4>
    <input class="search-bar" type="text" placeholder="Filter kills..." oninput="filterTable('tbl-kills',this.value)">
    <div class="table-wrap"><table id="tbl-kills">
      <thead><tr><th>Timestamp</th><th>GM Character</th><th>Target Name</th><th>Target ID</th></tr></thead>
      <tbody>{kill_log_html or '<tr><td colspan="4" style="color:#888;text-align:center">No kill records</td></tr>'}</tbody>
    </table></div>
    <h4 style="margin-top:1.5rem">Full Items Given Log</h4>
    <input class="search-bar" type="text" placeholder="Filter items..." oninput="filterTable('tbl-items',this.value)">
    <div class="table-wrap"><table id="tbl-items">
      <thead><tr><th>Timestamp</th><th>GM Character</th><th>Item ID</th><th>Item Name</th><th>Rarity</th><th>Qty</th><th>Command</th></tr></thead>
      <tbody>{item_log_html or '<tr><td colspan="7" style="color:#888;text-align:center">No item records</td></tr>'}</tbody>
    </table></div>
    <h4 style="margin-top:1.5rem">&#128128; Death Log <span class="sub">({len(deaths_raw)} records)</span></h4>
    <div class="table-wrap"><table id="tbl-deaths">
      <thead><tr><th>Timestamp</th><th>Character</th><th>Level</th><th>Killed By</th><th>Zone</th></tr></thead>
      <tbody>{deaths_html}</tbody>
    </table></div>
  </div>

  <div class="tab-pane" id="all">
    <input class="search-bar" type="text" placeholder="Search commands, items, characters..." oninput="filterTable('tbl-all',this.value)">
    <div class="table-wrap"><table id="tbl-all">
      <thead><tr><th>Timestamp</th><th>Character</th><th>Account</th><th>Category</th><th>Command</th><th>Item / Target</th></tr></thead>
      <tbody>{all_rows_html}</tbody>
    </table></div>
  </div>

  <div class="tab-pane" id="characters">
    <h3>&#128100; All Characters <span class="sub">({len(chars_raw)} total)</span></h3>
    <input class="search-bar" type="text" placeholder="Filter characters..." oninput="filterTable('tbl-chars',this.value)">
    <div class="table-wrap"><table id="tbl-chars">
      <thead><tr><th>Name</th><th>Class</th><th>Level</th><th>Race</th><th>Account</th><th>Guild</th><th>Last Seen</th></tr></thead>
      <tbody>{chars_html}</tbody>
    </table></div>
    <h3 style="margin-top:1.5rem">&#128200; Level Progression</h3>
    <div id="level-content">
      <div class="chart-wrap" style="margin-bottom:1.5rem"><canvas id="chart-levels" height="100"></canvas></div>
      <div class="section-head" style="margin-bottom:.75rem">&#9203; Leveling Speed by Character
        <span style="font-size:.75rem;font-weight:400;margin-left:.75rem;opacity:.7">Each cell = one level gained &nbsp;|&nbsp; Red = fastest &nbsp;&#8594;&nbsp; Green = slowest</span>
      </div>
      <div id="level-heatmap" style="padding:.75rem;background:var(--bg2);border:1px solid var(--border);border-radius:0 0 6px 6px">
        <span style="color:var(--text3);font-size:.85rem">Loading...</span>
      </div>
    </div>
  </div>

  <div class="tab-pane" id="spire-log">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1rem;flex-wrap:wrap;gap:.5rem">
      <h3 style="margin:0">&#128220; Spire Activity Log</h3>
      <div style="display:flex;gap:.5rem;align-items:center;flex-wrap:wrap">
        <input id="spire-user-filter" type="text" class="search-bar"
               style="width:140px;margin:0" placeholder="User...">
        <button onclick="loadSpireLog(document.getElementById('spire-user-filter').value)"
                style="background:var(--accent);color:#fff;border:none;border-radius:6px;
                       padding:.4rem 1rem;cursor:pointer;font-size:.82rem;font-weight:600">
          &#8635; Load / Refresh
        </button>
        <button onclick="loadSpireLog('')"
                style="background:var(--bg3);color:var(--text2);border:1px solid var(--border);
                       border-radius:6px;padding:.4rem .8rem;cursor:pointer;font-size:.82rem">
          Show All
        </button>
      </div>
    </div>
    <div id="spire-log-status" style="font-size:.8rem;color:var(--text3);margin-bottom:.6rem"></div>
    <input class="search-bar" type="text" placeholder="Filter table..."
           oninput="filterTable('tbl-spire-log',this.value)">
    <div class="table-wrap"><table id="tbl-spire-log">
      <thead><tr>
        <th>Timestamp</th><th>User</th><th>Action</th><th>Entity</th><th>Detail</th>
      </tr></thead>
      <tbody id="spire-log-body">
        <tr><td colspan="5" style="text-align:center;color:var(--text3);padding:2rem">
          Click <strong>Load / Refresh</strong> to fetch Spire activity for kphill.
        </td></tr>
      </tbody>
    </table></div>
  </div>

  {cat_panes}
  {char_panes}
</div>
<script>
function showTab(id,btn){{
  document.querySelectorAll('.tab-pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  if(btn) btn.classList.add('active');
}}
function filterTable(id,q){{
  const lq=q.toLowerCase();
  document.getElementById(id).querySelectorAll('tbody tr').forEach(r=>{{
    r.style.display=r.textContent.toLowerCase().includes(lq)?'':'none';
  }});
}}

const TEAL='#22d3a0', AMBER='#fbbf24', CHARCOAL='#131620', GRAY='#1e2235';

new Chart(document.getElementById('chart-daily'),{{
  type:'bar',
  data:{{
    labels:{trend_labels},
    datasets:[{{label:'Commands',data:{trend_data},backgroundColor:'#22d3a0aa',borderColor:'#22d3a0',borderWidth:1,borderRadius:3}}]
  }},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:20,maxRotation:45,color:'#4a5070'}},grid:{{color:'#1e2235'}}}},y:{{beginAtZero:true,ticks:{{color:'#4a5070'}},grid:{{color:'#1e2235'}}}}}}}}
}});

new Chart(document.getElementById('chart-hourly'),{{
  type:'bar',
  data:{{
    labels:{hourly_labels},
    datasets:[{{label:'Commands',data:{hourly_json},backgroundColor:'#22d3a0aa',borderColor:'#22d3a0',borderWidth:1,borderRadius:3}}]
  }},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{color:'#4a5070'}},grid:{{color:'#1e2235'}}}},y:{{beginAtZero:true,ticks:{{color:'#4a5070'}},grid:{{color:'#1e2235'}}}}}}}}
}});

new Chart(document.getElementById('chart-types'),{{
  type:'doughnut',
  data:{{
    labels:{cat_labels},
    datasets:[{{data:{cat_data},backgroundColor:['#22d3a0','#fbbf24','#f87171','#c084fc','#38bdf8','#4ade80','#60a5fa','#e879f9','#fb923c','#6b7280']}}]
  }},
  options:{{responsive:true,plugins:{{legend:{{position:'right',labels:{{color:'#7a88b0'}}}}}}}}
}});


// Commands data for client-side new-since-last-view
const CMDS={cmds_js};
const GENERATED="{generated}";

function catBadge(cat){{
  return `<span class="cat-${{cat.toLowerCase().replace(/ /g,'-')}}">${{cat}}</span>`;
}}

function fmtCmd(raw,cat){{
  if(!raw) return '';
  const _e=s=>s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  const gi=raw.match(/^(#(?:giveitem|summonitem)) +([0-9]+)/i);
  if(gi) return _e(gi[1])+' <span class="item-link" onclick="lookupItem('+gi[2]+')" title="Item #'+gi[2]+'">'+gi[2]+'</span>'+_e(raw.slice(gi[0].length));
  const m=raw.match(/^(#(?:buff|castspell|discipline|ae|aoe)) +([0-9]+)(.*)?$/i);
  if(m) return _e(m[1])+' <span class="spell-link" onclick="lookupSpell('+m[2]+')" title="Spell #'+m[2]+'">'+m[2]+'</span>'+_e(m[3]||'');
  if(cat==='Buff'){{
    const nm=raw.match(/([0-9]+)/);
    if(nm){{
      const id=nm[1],i=nm.index;
      return _e(raw.slice(0,i))+'<span class="spell-link" onclick="lookupSpell('+id+')" title="Spell #'+id+'">'+id+'</span>'+_e(raw.slice(i+id.length));
    }}
  }}
  return _e(raw);
}}


function renderNewCmds(lastView){{
  const valEl  = document.getElementById('new-count-val');
  const subEl  = document.getElementById('new-count-sub');
  const navBtn = document.getElementById('new-nav-btn');
  const navCount = document.getElementById('new-nav-count');
  const body   = document.getElementById('new-cmds-body');
  const sub    = document.getElementById('new-cmds-sub');

  // No baseline — treat all commands as new
  const newer = lastView ? CMDS.filter(r => r[0] > lastView) : CMDS;
  const ts    = lastView ? new Date(lastView).toLocaleString() : 'beginning';
  valEl.textContent = newer.length;
  subEl.textContent = lastView ? 'since ' + ts : 'all time — click Mark all seen to set baseline';
  subEl.textContent = 'since ' + ts;
  if(sub) sub.textContent = lastView ? 'since ' + ts + ' — ' + newer.length + ' commands' : newer.length + ' total commands';

  if(newer.length > 0){{
    navBtn.style.display = 'inline-flex';
    navCount.textContent = newer.length;
  }}

  if(body){{
    body.innerHTML = newer.length === 0
      ? '<tr><td colspan="6" style="text-align:center;color:var(--text3)">No new commands since last clear</td></tr>'
      : newer.map(r => `<tr><td>${{r[0]}}</td><td>${{r[1]}}</td><td>${{r[2]}}</td><td>${{catBadge(r[3])}}</td><td class="cmd">${{fmtCmd(r[4],r[3])}}</td><td class="${{r[6]}}">${{r[5]}}</td></tr>`).join('');
  }}
}}

function loadClearTime(){{
  fetch('/clear_time.json?_=' + Date.now())
    .then(r => r.json())
    .then(d => renderNewCmds(d.cleared_at || null))
    .catch(() => renderNewCmds(null));
}}

function clearNewCmds(){{
  fetch('/clear', {{method:'POST'}})
    .then(r => r.json())
    .then(d => renderNewCmds(d.cleared_at))
    .catch(() => {{
      // fallback to localStorage if server unreachable
      localStorage.setItem('gm-last-view', GENERATED);
      renderNewCmds(GENERATED);
    }});
}}

// Run on load
loadClearTime();

function toggleTheme(){{
  const html=document.documentElement;
  const btn=document.getElementById('theme-toggle');
  if(html.getAttribute('data-theme')==='light'){{
    html.removeAttribute('data-theme');
    btn.textContent='\u2600 Light';
    localStorage.setItem('gm-theme','dark');
  }}else{{
    html.setAttribute('data-theme','light');
    btn.textContent='\U0001f319 Dark';
    localStorage.setItem('gm-theme','light');
  }}
}}
(function(){{
  const saved=localStorage.getItem('gm-theme');
  const btn=document.getElementById('theme-toggle');
  if(saved==='light'){{
    document.documentElement.setAttribute('data-theme','light');
    btn.textContent='\U0001f319 Dark';
  }}
}})();

// Spire API live polling
function updateSpire(){{
  // System health
  fetch('/spire/api/v1/admin/system/resource-usage-summary?_='+Date.now())
    .then(r=>r.json()).then(d=>{{
      const cpu=d.cpu?d.cpu.toFixed(1):'—';
      const ram=d.memory?d.memory.usedPercent.toFixed(1):'—';
      const cpuEl=document.getElementById('sb-cpu');
      const ramEl=document.getElementById('sb-ram');
      const cpuBar=document.getElementById('sb-cpu-bar');
      const ramBar=document.getElementById('sb-ram-bar');
      if(cpuEl) cpuEl.textContent=cpu+'%';
      if(ramEl) ramEl.textContent=ram+'%';
      if(cpuBar) cpuBar.style.width=Math.min(d.cpu||0,100)+'%';
      if(ramBar) ramBar.style.width=Math.min(d.memory?d.memory.usedPercent:0,100)+'%';
      if(cpuBar) cpuBar.style.background=(d.cpu||0)>80?'#f87171':(d.cpu||0)>60?'#fbbf24':'#22d3a0';
      if(ramBar) ramBar.style.background=(d.memory&&d.memory.usedPercent>85)?'#f87171':'#60a5fa';
    }}).catch(()=>{{}});

  // Dashboard stats
  fetch('/spire/api/v1/eqemuserver/dashboard-stats?_='+Date.now())
    .then(r=>r.json()).then(d=>{{
      const s=id=>document.getElementById(id);
      if(s('sv-accounts')&&d.accounts!=null) s('sv-accounts').textContent=d.accounts;
      if(s('sv-characters')&&d.characters!=null) s('sv-characters').textContent=d.characters;
      if(s('sv-guilds')&&d.guilds!=null) s('sv-guilds').textContent=d.guilds;
    }}).catch(()=>{{}});

  fetch('/spire/api/v1/eqemuserver/server-stats?_='+Date.now())
    .then(r=>r.json()).then(d=>{{
      document.getElementById('sb-server').textContent=d.server_name||'—';
      document.getElementById('sb-uptime').textContent=(d.uptime||'').replace('Worldserver Uptime | ','').replace(', and',' and');
      document.getElementById('sb-players').textContent=d.players_online??'—';
      document.getElementById('sb-zones').textContent=d.zone_count??'—';
      const procs={{}};(d.main_process_stats||[]).forEach(p=>{{procs[p.name]=p;}});
      ['world','ucs'].forEach(n=>{{const dot=document.getElementById('sb-'+n+'-dot');if(dot)dot.className='proc-dot '+(procs[n]&&procs[n].pid>0?'proc-ok':'proc-down');}});
      const zd=document.getElementById('sb-zone-dot');if(zd)zd.className='proc-dot '+(d.zone_count>0?'proc-ok':'proc-down');
      const upd=document.getElementById('sb-updated');if(upd)upd.textContent='Updated '+new Date().toLocaleTimeString();
    }}).catch(()=>{{}});
  fetch('/spire/api/v1/eqemuserver/client-list?_='+Date.now())
    .then(r=>r.json()).then(d=>{{
      const clients=d.data||[];
      const onlineEl=document.getElementById('online-count');
      const subEl=document.getElementById('online-pane-count');
      const pane=document.getElementById('online-pane');
      if(onlineEl) onlineEl.textContent=clients.length;
      if(subEl) subEl.textContent=clients.length?'('+clients.length+' connected)':'';
      // Zone breakdown
      const zoneBreak=document.getElementById('zone-breakdown');
      const zoneChips=document.getElementById('zone-chips');
      const inGame=clients.filter(c=>c.online===1);
      if(zoneBreak){{
        if(inGame.length===0){{
          zoneBreak.style.display='none';
        }}else{{
          zoneBreak.style.display='';
          const zones={{}};
          inGame.forEach(c=>{{
            const z=(c.server&&(c.server.zone||c.server.zone_name||c.server.short_name))||'Unknown Zone';
            if(!zones[z])zones[z]=[];
            zones[z].push(c.name||(c.account_name||'?'));
          }});
          if(zoneChips) zoneChips.innerHTML=Object.entries(zones).map(([z,pls])=>
            '<span class="zone-chip"><span style="color:var(--accent);font-weight:600">'+z+'</span>'
            +'<span style="color:var(--text3)">'+pls.join(', ')+'</span></span>'
          ).join('');
        }}
      }}

      if(pane){{
        if(clients.length===0){{
          pane.innerHTML='<span style="color:var(--text3);font-size:.85rem">No players currently connected</span>';
        }}else{{
          // Deduplicate by account+name combo
          const seen=new Set();
          const unique=clients.filter(c=>{{
            const key=(c.account_name||'')+'|'+(c.name||'');
            if(seen.has(key)) return false;
            seen.add(key); return true;
          }});
          pane.innerHTML=unique.map(c=>
            '<div style="display:inline-flex;align-items:center;gap:.5rem;background:var(--bg3);border:1px solid '+(c.online===1?'#22d3a055':'var(--border)')+';border-radius:6px;padding:.4rem .8rem;margin:.25rem .25rem .25rem 0;font-size:.85rem">'
            +'<span style="width:8px;height:8px;border-radius:50%;background:'+(c.online===1?'#22d3a0':'#4a5070')+';'+(c.online===1?'box-shadow:0 0 5px #22d3a0;':'')+'flex-shrink:0"></span>'
            +'<strong style="color:'+(c.admin>0?'#fb923c':'var(--accent)')+';">'+(c.name||'(char select)')+'</strong>'
            +'<span style="color:var(--text2);font-size:.8rem"> '+c.account_name+'</span>'
            +(c.admin>0?'<span class="badge cat-gm-toggle" style="font-size:.7rem;padding:.1rem .4rem">GM</span>':'')
            +'<span style="color:var(--text3);font-size:.75rem"> '+(c.online===1?'In Game':'Char Select')+'</span>'
            +'</div>'
          ).join('');
        }}
      }}
    }}).catch(()=>{{}});
}}
updateSpire();setInterval(updateSpire,30000);

// ---- Activity Heatmap & Session Detection ----
function buildActivityHeatmap(cmds){{
  const days=['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
  const grid={{}};
  days.forEach(d=>{{ grid[d]={{}}; for(let h=0;h<24;h++) grid[d][h]=0; }});
  cmds.forEach(c=>{{
    const ts=c.timestamp||c.ts||'';
    if(!ts) return;
    const dt=new Date(ts.replace(' ','T'));
    if(isNaN(dt)) return;
    const day=days[dt.getDay()];
    const hr=dt.getHours();
    grid[day][hr]=(grid[day][hr]||0)+1;
  }});
  const allVals=days.flatMap(d=>Object.values(grid[d]));
  const maxV=Math.max(...allVals,1);

  let html='<div style="overflow-x:auto"><table style="border-collapse:collapse;font-size:.72rem;width:100%">';
  html+='<thead><tr><th style="padding:.3rem .5rem;color:var(--text3)">Day</th>';
  for(let h=0;h<24;h++) html+=`<th style="padding:.2rem;color:var(--text3);text-align:center;min-width:28px">${{h.toString().padStart(2,'0')}}</th>`;
  html+='</tr></thead><tbody>';
  days.forEach(day=>{{
    html+=`<tr><td style="padding:.3rem .6rem;color:var(--text2);font-weight:600;white-space:nowrap">${{day}}</td>`;
    for(let h=0;h<24;h++){{
      const v=grid[day][h];
      const t=v/maxV;
      const alpha=v===0?0.05:0.15+t*0.85;
      const bg=`rgba(34,211,160,${{alpha.toFixed(2)}})`;
      const tc=t>0.5?'#111':'var(--text2)';
      html+=`<td title="${{day}} ${{h.toString().padStart(2,'0')}}:00 — ${{v}} commands" style="background:${{bg}};color:${{tc}};text-align:center;padding:.25rem .2rem;border-radius:3px;cursor:default">${{v||''}}</td>`;
    }}
    html+='</tr>';
  }});
  html+='</tbody></table></div>';
  html+=`<div style="margin-top:.5rem;display:flex;align-items:center;gap:.5rem;font-size:.75rem;color:var(--text3)">
    <span>Low</span>
    <div style="width:120px;height:10px;border-radius:4px;background:linear-gradient(to right,rgba(34,211,160,.15),rgba(34,211,160,1))"></div>
    <span>High</span>
  </div>`;
  return html;
}}

function buildSessionTable(cmds){{
  if(!cmds.length) return '<p style="color:var(--text3)">No command data.</p>';
  const sorted=[...cmds].sort((a,b)=>(a.timestamp||'').localeCompare(b.timestamp||''));
  const sessions=[];
  let cur=null;
  sorted.forEach(c=>{{
    const ts=new Date((c.timestamp||'').replace(' ','T'));
    if(isNaN(ts)) return;
    if(!cur||((ts-cur.lastTs)/60000>30)){{
      if(cur) sessions.push(cur);
      cur={{start:ts,lastTs:ts,cmds:1,chars:new Set([c.character_name])}};
    }} else {{
      cur.lastTs=ts; cur.cmds++; cur.chars.add(c.character_name);
    }}
  }});
  if(cur) sessions.push(cur);
  sessions.sort((a,b)=>b.start-a.start);
  let html='<div class="table-wrap"><table><thead><tr><th>Session Start</th><th>Duration</th><th>Commands</th><th>Characters</th></tr></thead><tbody>';
  sessions.forEach(s=>{{
    const dur=Math.round((s.lastTs-s.start)/60000);
    const durStr=dur<60?dur+'m':Math.floor(dur/60)+'h '+(dur%60)+'m';
    const chars=[...s.chars].join(', ');
    const dateStr=s.start.toISOString().replace('T',' ').slice(0,16);
    html+=`<tr><td style="color:var(--text3);font-size:.82rem">${{dateStr}}</td>
      <td style="text-align:center;font-weight:600;color:var(--accent)">${{durStr}}</td>
      <td style="text-align:center">${{s.cmds}}</td>
      <td style="color:var(--text2);font-size:.82rem">${{chars}}</td></tr>`;
  }});
  html+='</tbody></table></div>';
  return html;
}}

// ── Spell lookup ─────────────────────────────────────────────────────────────
const _spellCache = {{}};

const EQ_TARGET_TYPES = {{
  0:'Self',1:'Pet',2:'Group v1',3:'Targeted',4:'AE Target',5:'Nearby Objects',
  6:'Self Group',7:'Group v2',8:'Directional',9:'AE',10:'All',11:'Target-AE',
  12:'Undead',13:'Animals',14:'AE Undead',16:'AE Target',18:'Targeted-AE',
  24:'Target Ring',32:'Beam',40:'AE Caster',41:'PB AE',42:'Target-AE Group',
  43:'Frontal Cone',44:'Circular Ring',45:'Directional Targetted'
}};
const EQ_RESIST_TYPES = {{
  0:'Unresistable',1:'Magic',2:'Fire',3:'Cold',4:'Poison',5:'Disease',
  6:'Chromatic',7:'Prismatic',8:'Physical',9:'Corruption'
}};
const EQ_SPELL_CLASSES = ['WAR','CLR','PAL','RNG','SHD','DRU','MNK','BRD','ROG','SHM','NEC','WIZ','MAG','ENC','BST','BER'];

function lookupSpellByName(name, btnId){{
  const status = btnId ? document.getElementById(btnId+'-status') : null;
  if(status) status.textContent = 'Searching...';
  // Build search candidates: full name, then first-4-words, first-3-words
  const words = name.trim().split(/ +/);
  const candidates = [...new Set([
    name,
    words.slice(0,4).join(' '),
    words.slice(0,3).join(' '),
    words.slice(0,2).join(' ')
  ].filter(s=>s.length>2))];

  function tryCandidate(i){{
    if(i>=candidates.length){{
      if(status) status.textContent='Spell not found: '+name;
      return;
    }}
    fetch('/spell-search?name='+encodeURIComponent(candidates[i])+'&_='+Date.now())
      .then(r=>{{if(!r.ok) throw new Error('HTTP '+r.status); return r.json();}})
      .then(rows=>{{
        const nl=name.toLowerCase();
        const match=rows.find(s=>(s.name||'').toLowerCase()===nl)
                   ||rows.find(s=>nl.startsWith((s.name||'').toLowerCase().slice(0,6)))
                   ||rows[0];
        if(!match){{tryCandidate(i+1);return;}}
        if(status) status.textContent='';
        closeItemModal();
        renderSpellCard(document.getElementById('spell-card-content'),match);
        document.getElementById('spell-modal').classList.add('open');
      }})
      .catch(e=>{{if(status) status.textContent='Error: '+e.message;}});
  }}
  tryCandidate(0);
}}


function closeSpellModal(){{
  document.getElementById('spell-modal').classList.remove('open');
}}

function lookupSpell(spellId){{
  const modal = document.getElementById('spell-modal');
  const el    = document.getElementById('spell-card-content');
  modal.classList.add('open');
  if(_spellCache[spellId]){{renderSpellCard(el,_spellCache[spellId]);return;}}
  el.innerHTML =
    '<button onclick="closeSpellModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button>'
    +'<div style="text-align:center;padding:2.5rem;color:var(--text3)">Loading spell #'+spellId+'...</div>';
  // Query DB directly — Spire has no spell-by-id REST endpoint on this version
  fetch('/spell-search?id='+spellId+'&_='+Date.now())
    .then(r=>{{if(!r.ok)throw new Error(r.status);return r.json();}})
    .then(rows=>{{
      if(!rows||!rows.length) throw new Error('not found');
      const d=rows[0];
      _spellCache[spellId]=d;
      closeItemModal();
      renderSpellCard(el,d);
    }})
    .catch(e=>{{
      el.innerHTML =
        '<button onclick="closeSpellModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button>'
        +'<div style="color:#f87171;padding:1.5rem;text-align:center">Spell #'+spellId+' not found: '+e.message+'</div>';
    }});
}}

// EQEmu Spell Proc Ability (SPA) effect ID → human-readable name
// Source: EQEmu common/spdat.h
const EQ_SPA_NAMES = {{
  0:'Hit Points', 1:'Armor Class', 2:'ATK', 3:'Movement Speed',
  4:'STR', 5:'DEX', 6:'AGI', 7:'STA', 8:'INT', 9:'WIS', 10:'CHA',
  11:'Haste', 12:'Invis vs Undead', 13:'See Invisible', 14:'Water Breathing',
  15:'Current Mana', 16:'Instant HP', 17:'NPC Frenzy', 18:'Identify',
  19:'Unblur', 20:'Charm', 21:'Fear', 22:'Stamina', 23:'Bind Affinity',
  24:'Gate', 25:'Cancel Magic (AE)', 26:'Inhibit Melee', 27:'Fear (2)',
  28:'Mesmerize', 29:'Summon Item', 30:'Summon Pet', 31:'Confuse',
  32:'Disease Counter', 33:'Poison Counter', 34:'Detect Hostility',
  35:'Dispel Magic', 36:'Invisibility', 37:'See Invisible (v2)',
  38:'Pacify', 39:'Ultravision', 40:'Add Faction', 41:'Stun',
  42:'Calm (AE)', 43:'Fear (3)', 44:'Feign Death', 45:'Summon Skeleton',
  46:'Teleport', 47:'Translocate', 48:'HP Regen', 49:'Max HP',
  50:'Resurrect', 51:'Summon Companion', 52:'Sense Summoned', 53:'Levitate',
  54:'Illusion', 55:'Damage Shield', 56:'Transfer Summoned', 57:'Divine Aura',
  58:'Melee Slow', 59:'Reverse Gravity', 60:'Damage Increase', 61:'Cancel Magic',
  62:'Lull', 63:'Strike Through', 64:'Silence', 65:'Max Mana', 66:'Bard Haste',
  67:'Root', 68:'HP Regen (2)', 69:'Absorb Magic', 70:'Inc Spell Damage',
  71:'Reduce Spell Mana', 72:'Shadow Step', 73:'Berserk', 74:'AE Mesmerize',
  75:'Rampage', 76:'AE Taunt', 77:'Flesh to Bone', 78:'Prepade (obsolete)',
  79:'Rezz', 80:'Summon Horse', 81:'Agro Reduce', 82:'Extreme Agro',
  83:'Summon Familiar', 84:'Summon Item Into Bag', 85:'Illusion: Target',
  86:'Mass Group Buff', 87:'AE Cancel Magic', 88:'Memory Blur',
  89:'Charm (2)', 90:'Size', 91:'Cloak', 92:'Add Hate', 93:'Stop Rain',
  94:'Snare', 95:'Portal (Translocate)', 96:'Alter NPC Level',
  97:'Summon Weapon', 98:'Self Resurrect', 99:'Silence (2)',
  100:'Sum Planar Familiar', 101:'Increase Pet Power', 102:'Weak Stun',
  103:'Melee Increase', 104:'Sanctification', 105:'Absorb Damage',
  106:'AE Undead', 107:'Remove Summon', 108:'Target Vision',
  109:'Fade', 110:'Stun (2)', 111:'Resurrect All', 112:'AE Charm',
  113:'Strip Virtual Pet Buff', 114:'Disintegrate', 115:'Invis vs Animals',
  116:'AE Dispel Magic', 117:'Mana Burn', 118:'Distraction',
  119:'Disguise', 120:'Melee Proc', 121:'HP/Mana Regen',
  122:'Levitate (2)', 123:'Add Class to Pet', 124:'Dispel Beneficial',
  125:'PvP Damage Mitigation', 126:'Spell Damage Mitigation',
  127:'Return Target HP', 128:'Gate 2', 129:'Translocate to Anchor',
  130:'Translocate to Bind', 131:'Arch Teleport', 132:'Reduce Timer',
  133:'Limit: Max Level', 134:'Limit: Resist Type', 135:'Limit: Target Type',
  136:'Limit: Effect', 137:'Limit: SpellType', 138:'Limit: Spell',
  139:'Limit: Min Duration', 140:'Limit: Instant Spells Only',
  141:'Limit: Min Level', 142:'Limit: Min Mana', 143:'Ranger Attack',
  144:'Rear Arc', 145:'Eye of Zomm', 146:'Recall Pet', 147:'Pet Damage',
  148:'Stacking: Block', 149:'Stacking: Overwrite', 150:'Death Save',
  151:'Suspended Minion', 152:'Tribute', 153:'Give Double Riposte',
  154:'Projectile', 155:'Pet Stacking', 156:'Accuracy',
  157:'Headshot', 158:'Pet Crit Melee', 159:'Slay Undead',
  160:'Bonus Attack', 161:'Frontal Arc', 162:'Melee Lifetap',
  163:'Stamina (2)', 164:'NPC Faction', 165:'Illusion: Friendly',
  166:'Pacify (2)', 167:'Spell Casting Mastery', 168:'Block Next Spell',
  169:'Illusion: Tree', 170:'Increase Damage', 171:'Reduce Damage',
  172:'AC (2)', 173:'Mana Conservation', 174:'Pet Discipline',
  175:'Return to Home', 176:'Healing Mastery', 177:'Avoidance',
  178:'Accuracy (2)', 179:'Damage Shield (2)', 180:'Body Type',
  181:'Modify Debuff Duration', 182:'Inc Spell Damage (v2)',
  183:'Inc Healing (v2)', 184:'Reverse Damage Shield',
  185:'Reduce Weight', 186:'Block Melee', 187:'Block Spell',
  188:'Reduce Mana Regen', 189:'Damage (2)', 190:'HP Regen Cap',
  191:'Mana Regen Cap', 192:'Stasis', 193:'Aging',
  194:'Defense Bonus', 195:'Melee Delay', 196:'Mod Rod',
  197:'Packrat', 198:'Block Behind', 199:'Endurance Regen',
  200:'Endurance', 201:'Endurance Regen (2)', 202:'Stun Resist',
  203:'Strikethrough (2)', 204:'Skill Increase',
  205:'Pet Innate Run Speed', 206:'Corpse Bomb',
  207:'Masquerade', 208:'Siphon HP', 209:'Target of Target',
  210:'Limit: Mana Min', 211:'Gravity Effect', 212:'Display Spell',
  213:'Increase Potion Belt', 214:'Increase Bandolier',
  215:'Pet Talent', 216:'Absorb Partial Melee',
  217:'Total HP Increase', 218:'Increase Clairvoyance',
  219:'Foraging', 220:'Innate Run Speed Cap', 221:'HP Percent',
  222:'Pet Affinity', 223:'Survivability', 224:'Origin',
  225:'Spell Crit Chance', 226:'Improve Parry', 227:'Balance Party HP',
  228:'Conversion (Mana to HP)', 229:'Conversion (HP to Mana)',
  230:'Limit: Focus Target', 231:'Increase Target Cap',
  232:'Faction (v3)', 233:'Decrease Skill', 234:'Limit: Instant Type',
  235:'Decrease Curse Counter', 236:'Spell Recast Time',
  237:'Limit: SpellGroup', 238:'Limit: Max Duration',
  239:'Limit: Min HP', 240:'Limit: Min Mana (2)',
  241:'Focus Petrification', 242:'Spell Proc',
  243:'Reduce Timer (2)', 244:'Limit: Classes',
  245:'Double Attack', 246:'Pet Flurry',
  247:'Reduce Hate', 248:'Gate to Home City',
  249:'Defensive Proc', 250:'Pet Hp Max',
  251:'Reflect Spell', 252:'Toggle Aggro',
  253:'Add Skill Innate', 254:'Deactivate All Traps',
  255:'Learn Trap', 256:'Change Trigger Type',
  257:'Mute', 258:'NPC Melee Slow',
  259:'Spell Damage Shield', 260:'Fear (Break on Hit)',
  261:'Spell Cancel', 262:'Proc',
  263:'Force Rebuff', 264:'Affliction Haste',
  265:'Twincast', 266:'Servile',
  267:'Shroud of Stealth', 268:'Give Pet Hold',
  269:'Triple Backstab', 270:'Divine Save',
  271:'Remove Detrimental', 272:'Reduce Duration',
  273:'Limit: Min Level (2)', 274:'Limit: HP Percent',
  275:'Limit: Current Mana Percent', 276:'Apply Effect',
  277:'Reuse Timer', 278:'Limit: Type',
  279:'Buff Blocker', 280:'Increase PC Attack Speed',
  281:'Reduce NPC Melee Skill', 282:'Remove Beneficial',
  283:'Decrease Mana', 284:'Inc Spell Damage (v3)',
  285:'Limit: Max Level (2)', 286:'Limit: Min Mana',
  287:'Mana Shield', 288:'Rune',
  289:'Mana Shield (2)', 290:'AE Melee',
  291:'Harmonious Attack', 292:'HP to Endurance',
  293:'Endurance to HP', 294:'Improved Healing',
  295:'Limit: SpellCategory', 296:'Limit: Min Mana (3)',
  297:'Total HP Increase (2)', 298:'Decrease Skill Level',
  299:'HP at Level', 300:'Faction (v4)',
  301:'Summon Mount', 302:'Improved Damage (2)',
  303:'Reverse Sanctuary', 304:'Mag Increase',
  305:'Balance Party Mana', 306:'Limit: PC HP Percent',
  307:'Limit: PC Mana Percent', 308:'Increase Haste (v3)',
  309:'Reverse Damage Shield (2)', 310:'Critical Hit',
  311:'Crit Spell', 312:'Crit Heal', 313:'Crit DoT',
  314:'Crit Mez', 315:'Crit Nuke',  316:'Dodge',
  317:'Inc Crit Melee', 318:'Inc Crit Heal',
  319:'Inc Crit Nuke', 320:'Inc Crit DoT',
  321:'Limit: Min Level (3)', 322:'Limit: Max Mana',
  323:'Pct Mana HP', 324:'Absorb Spell Proc',
  325:'Limit: SpellGroup (2)', 326:'HC Absorb',
  327:'Inc Hits', 328:'Limit: CastTime',
  329:'Limit: Max CastTime', 330:'Limit: Min CastTime',
  331:'Limit: PC HP Pct', 332:'Mana Drain',
  333:'Limit: Level', 334:'Limit: Mana',
  335:'Skill Attack', 336:'Decrease Hits',
  337:'Limit: PC Mana Pct', 338:'Inc Weapon Dmg',
  339:'Limit: Max Mana (2)', 340:'Decrease Skill Attack',
  341:'Summon Corpse', 342:'Heal Rate',
  343:'Reverse Fly', 344:'No Break AE Sneak',
  345:'Slay (2)', 346:'Pet Max HP',
  347:'Merchant Modifier', 348:'Inc Spell Procs',
  349:'Learn Skill', 350:'Add Pet Command',
  351:'Extended Target Window', 352:'Mana Absorb',
  353:'Focus Duration Max', 354:'Focus Detrimental',
  355:'Focus Beneficial', 356:'Fear',
  357:'Cancel Aggro', 358:'Bind (2)',
  359:'Gate (2)', 360:'Divine Save (2)',
  361:'Spell Damage Taken', 362:'Melee ATK Damage',
  363:'Worn Attack', 364:'Damage Shield (3)',
  365:'Limit: Spell Cast Time Min', 366:'Limit: SpellType (2)',
  367:'Crit Melee Pct', 368:'Inc Spell Dmg (v4)',
  369:'Decrease Poison Counter', 370:'Decrease Disease Counter',
  371:'Decrease Curse Counter', 372:'Decrease Corruption Counter',
  373:'Increase Corruption Counter', 374:'Corruption Counter',
  375:'Focus Wpn Dmg', 376:'Focus Mana Regen',
  377:'Focus HP Regen', 378:'Focus Haste',
  379:'Focus Pet Haste', 380:'Focus Pet Proc',
  381:'Focus Pet Lifetap', 382:'Focus Pet Innate',
  383:'Focus Pet Endurance', 384:'Focus Pet Mana',
  385:'Focus Pet Stats', 386:'Focus Pet Haste (2)',
  387:'Focus Pet Regen', 388:'Focus Pet Accuracy',
  389:'Focus Pet Avoidance', 390:'Focus Pet Damage',
  391:'Focus Pet Crit', 392:'Focus Pet Mitigation',
  393:'Focus Spell Crit', 394:'Focus Healing',
  395:'Focus Spell Damage %', 396:'Focus Pet Armor',
  397:'Focus Effectiveness', 398:'HP Per Tick',
  399:'Mana Per Tick', 400:'Focus Beneficial',
  401:'Focus Detrimental (2)', 402:'Focus Healing (2)',
  403:'Focus Duration', 404:'Focus Spell Dmg (5)',
  405:'Focus Resist', 406:'Focus Recast Delay',
  407:'Focus Mana Cost', 408:'Focus Range',
  409:'Focus Healing (3)', 410:'Focus Spell Dmg (6)',
  411:'Focus Duration (2)', 412:'Focus Pet Dmg',
  413:'Decrease Mana Cost', 414:'Focus Endurance Cost',
  415:'Focus Healing (4)', 416:'Focus Spell Dmg (7)',
  417:'Summon to Corpse', 418:'Reclaim Energy',
  419:'Translocate (2)', 420:'Recall',
  421:'Spell Focus', 422:'Spell Focus (2)',
  423:'Mount', 424:'Unknown',
  425:'Limit: Spell List', 426:'Limit: SpellGroup (3)',
  427:'Limit: SpellSubGroup', 428:'Resist',
  429:'Focus Effectiveness (2)', 430:'Focus Effectiveness (3)',
  431:'Focus Effectiveness (4)', 432:'Focus Limit Level',
  433:'Focus Limit Resist', 434:'Focus Limit Target',
  435:'Focus Limit Effect', 436:'Focus Limit SpellType',
  437:'Focus Limit Spell', 438:'Focus Min Duration',
  439:'Focus Instant Only', 440:'Focus Min Level',
  441:'Focus Min Mana', 442:'Focus Target',
  443:'Focus Spell Group', 444:'Focus Max Duration',
  445:'Focus Max Mana', 446:'Focus Class',
  447:'Focus Cast Time Min', 448:'Focus Cast Time Max',
  449:'Focus Ignore Resist', 450:'Focus Hit Points',
  451:'Focus Beneficial (3)', 452:'Focus Detrimental (3)',
  453:'Focus Healing (5)', 454:'Focus Spell Dmg (8)',
  455:'Duration', 456:'HP (v2)',
  457:'Focus Spell (2)', 458:'Jumpy',
  459:'Caster Illusion', 460:'Group Illusion',
  461:'Inc Spell Dmg %', 462:'Pet Affinity (2)',
}};


function renderSpellCard(el,d){{
  // Normalise field names — DB returns casttime/effectid1, Spire returns cast_time/effect_id_1
  const _ct = d.cast_time     ??d.casttime     ??0;
  const _rt = d.recast_time   ??d.recasttime   ??0;
  const _rv = d.recovery_time ??d.RecoveryTime ??d.recoverytime??0;
  const _tt = d.target_type   ??d.targettype   ??d.TargetType  ??0;
  const _rs = d.resist_type   ??d.ResistType   ??d.resisttype  ??0;
  const _bd = d.buffduration  ??d.BuffDuration ??0;
  const _si = d.spell_icon    ??d.spellicon    ??d.SpellIcon   ??0;
  const _cl = d.classes       ??d.Classes      ??0;

  const castS   = _ct===0 ? 'Instant' : (_ct/1000).toFixed(1)+'s';
  const recastS = _rt>0   ? (_rt/1000).toFixed(1)+'s' : 'None';
  const recTimeS= _rv>0   ? (_rv/1000).toFixed(1)+'s' : 'None';
  const target  = EQ_TARGET_TYPES[_tt] || ('Type '+_tt);
  const resist  = EQ_RESIST_TYPES[_rs] || 'None';
  const durTicks= _bd;
  const durStr  = durTicks>0 ? durTicks+' ticks ('+(durTicks*6)+' sec)' : 'Instant / No buff';
  const classStr= _cl===65535 ? 'ALL'
    : EQ_SPELL_CLASSES.filter((_,i)=>_cl&(1<<i)).join(' ')||'None';

  // Collect non-zero spell effects — handle both effectid1 and effect_id_1 naming
  const effects=[];
  for(let i=1;i<=12;i++){{
    const eid=d['effect_id_'+i]??d['effectid'+i]??0;
    const ebv=d['effect_base_value'+i]??d['effect_base_value_'+i]??d['effectbasevalue'+i]??0;
    if(eid&&eid!==254&&eid!==0){{
      const elv=d['effect_limit_value'+i]??d['effect_limit_value_'+i]??0;
      const lbl=elv&&elv!==0?' [limit:'+elv+']':'';
      const ename=EQ_SPA_NAMES[eid]||('Effect ID '+eid);
      effects.push(ename+': '+ebv+lbl);
    }}
  }}

  el.innerHTML=
    '<button onclick="closeSpellModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer;padding:0;line-height:1">&#10005;</button>'
    +'<div style="font-size:1.05rem;font-weight:700;color:#a78bfa;margin-bottom:.2rem">'+d.name+'</div>'
    +'<div style="font-size:.75rem;color:var(--text3);margin-bottom:.85rem">Spell ID: '+d.id
      +(_si?'  &nbsp;&#127775; Icon '+_si:'')+'</div>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:.3rem .9rem;font-size:.82rem;color:var(--text2);margin-bottom:.75rem">'
      +'<div>Mana: <strong style="color:'+(d.mana>0?'#60a5fa':'var(--text)')+'">'+((d.mana)||0)+'</strong></div>'
      +'<div>Cast: <strong style="color:var(--text)">'+castS+'</strong></div>'
      +'<div>Recast: <strong style="color:var(--text)">'+recastS+'</strong></div>'
      +'<div>Recovery: <strong style="color:var(--text)">'+recTimeS+'</strong></div>'
      +'<div>Target: <strong style="color:var(--text)">'+target+'</strong></div>'
      +'<div>Resist: <strong style="color:#22d3a0">'+resist+'</strong></div>'
      +(d.range>0?'<div>Range: <strong style="color:var(--text)">'+d.range+' units</strong></div>':'')
      +(d.aoerange>0?'<div>AoE Range: <strong style="color:var(--text)">'+d.aoerange+' units</strong></div>':'')
    +'</div>'
    +'<div style="font-size:.78rem;margin-bottom:.6rem">'
      +'<span style="color:var(--text3)">Duration: </span>'
      +'<strong style="color:'+(durTicks>0?'#fbbf24':'var(--text3)')+'">'+durStr+'</strong>'
    +'</div>'
    +(effects.length?'<div style="font-size:.75rem;color:var(--text3);margin-bottom:.5rem">Effects:<br>'
      +effects.map(e=>'<span style="color:var(--text2)">'+e+'</span>').join('<br>')+'</div>':'')
    +'<div style="font-size:.75rem;color:var(--text3)">Classes: <span style="color:var(--text2)">'+classStr+'</span></div>';
}}


// ── Spire audit log ───────────────────────────────────────────────────────────
window._spireLogLoaded = false;

function loadSpireLog(userFilter){{
  const body   = document.getElementById('spire-log-body');
  const status = document.getElementById('spire-log-status');
  const filter = (userFilter !== undefined ? userFilter
                : (document.getElementById('spire-user-filter').value || '')).toLowerCase().trim();

  body.innerHTML = '<tr><td colspan="5" style="text-align:center;color:var(--text3);padding:2rem">Loading Spire activity...</td></tr>';
  if(status) status.textContent = '';

  // Try the Spire audit-log endpoint; fall back to audit_log if needed
  fetch('/spire-user-log?_='+Date.now())
    .then(r=>{{if(!r.ok) throw new Error('HTTP '+r.status+' — endpoint may differ on this Spire version'); return r.json();}})
    .then(raw=>{{
      // Normalise: handle paginated {{data:[...]}} or flat array
      const all = Array.isArray(raw) ? raw
                : (raw.data || raw.audit_log || raw.entries || raw.logs || []);

      const rows = filter
        ? all.filter(e=>{{
            const u=(e.user_name||e.user||e.account||e.username||'').toLowerCase();
            return u.includes(filter);
          }})
        : all;

      if(status) status.textContent =
        rows.length+' entries'+(filter?' for "'+filter+'"':'')+
        ' — '+all.length+' total in log';

      if(!rows.length){{
        body.innerHTML='<tr><td colspan="5" style="text-align:center;color:var(--text3);padding:2rem">'+
          'No entries'+(filter?' matching "'+filter+'"':'')+
          ' — try <strong>Show All</strong> to see unfiltered results.</td></tr>';
        return;
      }}

      body.innerHTML = rows.map(e=>{{
        const ts     = e.created_at||e.timestamp||e.date||'';
        const user   = e.user_name||e.user||e.account||e.username||('User #'+(e.user_id||'?'));
        const action = e.event_name||e.event||e.action||e.type||e.operation||'?';
        const entity = [e.entity_type||e.resource_type||e.table_name||e.model||'',
                        e.entity_id  ||e.resource_id  ||'']
                       .filter(Boolean).join(' #') || '—';

        // Detail: prefer a diff-style view if from_values/to_values exist
        let detail = '—';
        if(e.from_values||e.to_values){{
          const from = e.from_values ? JSON.stringify(e.from_values).slice(0,150) : '';
          const to   = e.to_values   ? JSON.stringify(e.to_values  ).slice(0,150) : '';
          detail = (from?'<span style="color:#f87171">− '+from+'</span><br>':'')+
                   (to  ?'<span style="color:#4ade80">+ '+to  +'</span>':'');
        }} else if(e.detail||e.description||e.message){{
          detail = String(e.detail||e.description||e.message).slice(0,200);
        }} else if(e.data){{
          detail = JSON.stringify(e.data).slice(0,200);
        }}

        return '<tr>'+
          '<td style="white-space:nowrap;color:var(--text3);font-size:.8rem">'+ts+'</td>'+
          '<td style="color:var(--accent);font-weight:600">'+user+'</td>'+
          '<td><span class="badge cat-other" style="font-size:.72rem">'+action+'</span></td>'+
          '<td style="font-size:.82rem;color:var(--text2)">'+entity+'</td>'+
          '<td class="cmd" style="max-width:340px;font-size:.76rem;white-space:normal">'+detail+'</td>'+
        '</tr>';
      }}).join('');

      window._spireLogLoaded = true;
    }})
    .catch(e=>{{
      body.innerHTML='<tr><td colspan="5" style="padding:1.5rem;text-align:center">'+
        '<div style="color:#f87171;margin-bottom:.6rem">&#9888; '+e.message+'</div>'+
        '<div style="color:var(--text3);font-size:.8rem;line-height:1.6">'+
          'Common Spire audit log endpoints to check:<br>'+
          'Check that reports_server.py is running and DB is reachable.'+
        '</div>'+
      '</td></tr>';
    }});
}}


function initTrendsExtra(){{
  try{{
    const allCmds = window._allCmds||[];
    if(!allCmds.length){{
      document.getElementById('activity-heatmap').innerHTML='<span style="color:#f87171">No command data available</span>';
      return;
    }}
    const hm=document.getElementById('activity-heatmap');
    const st=document.getElementById('session-table');
    if(hm) hm.innerHTML=buildActivityHeatmap(allCmds);
    if(st) st.innerHTML=buildSessionTable(allCmds);
  }}catch(e){{
    const hm=document.getElementById('activity-heatmap');
    if(hm) hm.innerHTML='<span style="color:#f87171">Error: '+e.message+'</span>';
    console.error('initTrendsExtra error:',e);
  }}
}}

// Level progression chart
let _levelsLoaded = false;
const CHART_COLORS_LVL = ['#22d3a0','#fb923c','#60a5fa','#a78bfa','#fbbf24','#f87171','#4ade80','#e879f9'];

function fmtElapsed(m){{
  if(m===null||m===undefined) return '—';
  if(m<1) return '<1m';
  if(m<60) return Math.round(m)+'m';
  const h=Math.floor(m/60),rem=Math.round(m%60);
  return rem>0?h+'h '+rem+'m':h+'h';
}}

function elapsedColor(val, minV, maxV){{
  // Red = fastest (shortest time), Green = slowest (longest time)
  if(val===null||val===undefined||minV===maxV) return '#888';
  const t = Math.max(0, Math.min(1, (val - minV) / (maxV - minV)));
  // Red(239,68,68) -> Yellow(251,191,36) -> Green(34,197,94)
  let r,g,b;
  if(t < 0.5){{
    const s = t * 2;
    r = Math.round(239 + (251-239)*s);
    g = Math.round(68  + (191-68)*s);
    b = Math.round(68  + (36-68)*s);
  }} else {{
    const s = (t - 0.5) * 2;
    r = Math.round(251 + (34-251)*s);
    g = Math.round(191 + (197-191)*s);
    b = Math.round(36  + (94-36)*s);
  }}
  return `rgb(${{r}},${{g}},${{b}})`;
}}

function loadLevelChart(){{
  if(_levelsLoaded) return;
  fetch('/levels?_='+Date.now())
    .then(r=>{{if(!r.ok) throw new Error('HTTP '+r.status); return r.json();}})
    .then(data=>{{
      _levelsLoaded = true;
      const names = Object.keys(data);
      if(!names.length){{
        const el=document.getElementById('level-content');
        if(el) el.innerHTML='<p style="color:var(--text3)">No level history found.</p>';
        return;
      }}

      // Collect all elapsed values to normalize color gradient
      const allElapsed = names.flatMap(n=>data[n].map(p=>p.elapsed_min)).filter(v=>v!==null&&v!==undefined);
      const minE = Math.min(...allElapsed);
      const maxE = Math.max(...allElapsed);

      // ---- Line chart ----
      const allTs=[...new Set(names.flatMap(n=>data[n].map(p=>p.ts)))].sort();
      const datasets=names.map((name,i)=>{{
        const pts=data[name];
        const vals=allTs.map(ts=>{{
          const match=pts.filter(p=>p.ts<=ts);
          return match.length?match[match.length-1].to_level:null;
        }});
        return{{label:name,data:vals,
          borderColor:CHART_COLORS_LVL[i%CHART_COLORS_LVL.length],
          backgroundColor:'transparent',tension:0,stepped:'after',
          pointRadius:2,borderWidth:2,spanGaps:false}};
      }});
      const ctx=document.getElementById('chart-levels');
      if(ctx&&typeof Chart!=='undefined'){{
        new Chart(ctx,{{
          type:'line',
          data:{{labels:allTs,datasets}},
          options:{{
            scales:{{
              x:{{ticks:{{color:'#888',maxTicksLimit:10,maxRotation:45}}}},
              y:{{min:1,ticks:{{color:'#888',stepSize:5}},
                 title:{{display:true,text:'Level',color:'#888'}}}}
            }},
            plugins:{{
              legend:{{labels:{{color:'var(--text2)',boxWidth:12}}}},
              tooltip:{{callbacks:{{
                title:i=>i[0].label,
                label:i=>i.dataset.label+': Lvl '+i.raw
              }}}}
            }},
            animation:false,responsive:true,maintainAspectRatio:true
          }}
        }});
      }}

      // ---- Heatmap: one row per character, one cell per level ----
      const heatmap=document.getElementById('level-heatmap');
      if(!heatmap) return;

      let html='';
      names.forEach((name,ni)=>{{
        const charColor=CHART_COLORS_LVL[ni%CHART_COLORS_LVL.length];
        const pts=data[name];
        // Build level map: to_level -> {{elapsed, jumped, ts}}
        const lvlMap={{}};
        pts.forEach(p=>{{
          // Track all events including losses for display
          const key=p.to_level+'_'+p.ts;
          lvlMap[key]={{elapsed:p.elapsed_min,jumped:p.jumped,ts:p.ts,from:p.from_level,lvl:p.to_level}};
        }});
        const levels=Object.values(lvlMap).sort((a,b)=>a.ts.localeCompare(b.ts)||a.lvl-b.lvl);
        const maxLvl=levels[levels.length-1]||1;

        html+=`<div style="margin-bottom:1.25rem">
          <div style="display:flex;align-items:center;gap:.6rem;margin-bottom:.4rem">
            <strong style="color:${{charColor}};min-width:100px;font-size:.9rem">${{name}}</strong>
            <span style="color:var(--text3);font-size:.75rem">Lvl 1 → ${{maxLvl}} &nbsp;|&nbsp; ${{levels.length}} gains</span>
          </div>
          <div style="display:flex;flex-wrap:wrap;gap:3px">`;

        levels.forEach(entry=>{{
          const lvl=entry.lvl;
          const d=entry;
          const isLoss=d.jumped<0;
          const isBigJump=d.jumped>3;
          const isGMSet=Math.abs(d.jumped)>10;
          // Only color normal gains; losses and GM sets get special treatment
          const bg=isLoss?'#374151':isGMSet?'#7c3aed':isBigJump?'#b45309':elapsedColor(d.elapsed,minE,maxE);
          const textColor=isLoss||isGMSet?'#e5e7eb':'#111';
          const border=isGMSet?'2px solid #a78bfa':isBigJump?'2px solid #fbbf24':isLoss?'1px solid #6b7280':'1px solid rgba(0,0,0,.2)';
          const jumpLabel=isLoss?'▼'+Math.abs(d.jumped):isBigJump?'+'+d.jumped:'';
          const tooltip=`Lvl ${{d.from}}→${{lvl}} | ${{d.ts}} | ${{fmtElapsed(d.elapsed)}}${{isLoss?' | LEVEL LOSS':isGMSet?' | GM SET':isBigJump?' | +'+d.jumped+' levels!':''}}`;
          html+=`<div title="${{tooltip}}" style="background:${{bg}};border:${{border}};border-radius:4px;padding:.2rem .35rem;font-size:.72rem;color:${{textColor}};font-weight:600;cursor:default;min-width:28px;text-align:center;line-height:1.4">
            ${{lvl}}${{jumpLabel?'<sup style="font-size:.58rem;font-weight:700">'+jumpLabel+'</sup>':''}}
          </div>`;
        }});

        html+='</div></div>';
      }});

      // Legend
      html+=`<div style="margin-top:.75rem;display:flex;flex-wrap:wrap;align-items:center;gap:.75rem;font-size:.75rem;color:var(--text3)">
        <span>Fast</span>
        <div style="width:120px;height:10px;border-radius:4px;background:linear-gradient(to right,rgb(239,68,68),rgb(251,191,36),rgb(34,197,94))"></div>
        <span>Slow</span>
        <span style="margin-left:.5rem;display:flex;align-items:center;gap:.3rem"><span style="background:#7c3aed;border-radius:3px;padding:.1rem .35rem;color:#e5e7eb">61</span> GM level set</span>
        <span style="display:flex;align-items:center;gap:.3rem"><span style="background:#b45309;border:2px solid #fbbf24;border-radius:3px;padding:.1rem .35rem;color:#111">55<sup>+4</sup></span> multi-level jump</span>
        <span style="display:flex;align-items:center;gap:.3rem"><span style="background:#374151;border-radius:3px;padding:.1rem .35rem;color:#e5e7eb">45<sup>▼1</sup></span> level loss</span>
      </div>`;

      heatmap.innerHTML=html;
    }}).catch(e=>{{
      const hm=document.getElementById('level-heatmap');
      if(hm) hm.innerHTML='<span style="color:#f87171;font-size:.85rem">&#9888; Failed to load level data: '+e.message+'<br>Make sure the latest reports_server.py is deployed and the container restarted.</span>';
      console.error('Level chart error:',e);
    }});
}}

// Level chart loads when Characters tab is clicked (see onclick below)


// Auto-refresh disabled


// ── Inventory lookup ─────────────────────────────────────────
const EQ_CLASS_NAMES={{1:'Warrior',2:'Cleric',3:'Paladin',4:'Ranger',5:'Shadow Knight',
  6:'Druid',7:'Monk',8:'Bard',9:'Rogue',10:'Shaman',11:'Necromancer',
  12:'Wizard',13:'Magician',14:'Enchanter',15:'Beastlord',16:'Berserker'}};
const EQ_SLOT_NAMES={{0:'Charm',1:'Ear',2:'Head',3:'Face',4:'Ear',5:'Neck',
  6:'Shoulder',7:'Arms',8:'Back',9:'L.Wrist',10:'R.Wrist',11:'Hands',
  12:'Primary',13:'Secondary',14:'L.Ring',15:'R.Ring',
  16:'Chest',17:'Legs',18:'Feet',19:'Waist',20:'P.Source',21:'Ammo'}};

const _invCache={{}};

function closeInvModal(){{
  document.getElementById('inv-modal').classList.remove('open');
}}

function lookupInventory(name,type){{
  const modal=document.getElementById('inv-modal');
  const el=document.getElementById('inv-card-content');
  modal.classList.add('open');
  const key=type+':'+name;
  if(_invCache[key]){{renderInvCard(el,_invCache[key]);return;}}
  el.innerHTML='<button onclick="closeInvModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button>'
    +'<div style="text-align:center;padding:2.5rem;color:var(--text3)">Loading inventory for '+name+'...</div>';
  fetch('/inventory/'+type+'/'+encodeURIComponent(name))
    .then(r=>{{if(!r.ok)throw new Error(r.status);return r.json();}})
    .then(d=>{{_invCache[key]=d;renderInvCard(el,d);}})
    .catch(e=>{{el.innerHTML='<button onclick="closeInvModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button>'
      +'<div style="color:#f87171;padding:1.5rem;text-align:center">Failed to load inventory: '+e.message+'</div>';}});
}}

function renderInvCard(el,data){{
  const c=data.character||{{}};
  const cls=EQ_CLASS_NAMES[c.class]||('Class '+c.class);
  const isBot=data.is_bot;
  const nameColor=isBot?'#a78bfa':'#fb923c';
  const slots=data.slots||[];

  let rows='';
  if(slots.length===0){{
    rows='<div style="color:var(--text3);font-style:italic;padding:1rem 0">No equipped items found</div>';
  }}else{{
    slots.forEach(s=>{{
      const slotLabel=EQ_SLOT_NAMES[s.slotid]||('Slot '+s.slotid);
      const iid=s.item_id_real||s.itemid||0;
      const nm=s.item_name||(iid?'Unknown #'+iid:'Empty');
      const hasItem=iid>0&&nm&&nm!=='Empty';
      const miniStats=[];
      if(s.ac>0) miniStats.push('AC '+s.ac);
      if(s.hp>0) miniStats.push('HP +'+s.hp);
      if(s.mana>0) miniStats.push('Mana +'+s.mana);
      if(s.damage>0) miniStats.push(s.damage+'/'+s.delay);
      const stats=miniStats.length?'<div class="inv-mini-stats">'+miniStats.join(' &nbsp;')+'</div>':'';
      rows+='<div class="inv-slot">'
        +'<span class="inv-slot-label">'+slotLabel+'</span>'
        +(hasItem
          ?'<div><span class="inv-item-name" onclick="lookupItem('+iid+')" style="color:'+(s.nodrop?'#60a5fa':s.magic?'#4ade80':'var(--text)')+'">'+nm+'</span>'+stats+'</div>'
          :'<span class="inv-empty">—</span>')
        +'</div>';
    }});
  }}

  el.innerHTML=
    '<button onclick="closeInvModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button>'
    +'<div style="font-size:1.05rem;font-weight:700;color:'+nameColor+';margin-bottom:.15rem">'+c.name+'</div>'
    +'<div style="font-size:.8rem;color:var(--text3);margin-bottom:1rem">'+cls+' &nbsp;Level '+c.level+(isBot?' <span style=\"background:#a78bfa22;color:#a78bfa;padding:.1rem .4rem;border-radius:3px;font-size:.7rem\">BOT</span>':'')+'</div>'
    +'<div>'+rows+'</div>';
}}

// ── Item lookup ──────────────────────────────────────────────
const _itemCache={{}};
const SLOT_NAMES=['Charm','Ear','Head','Face','Ear2','Neck','Shoulder','Arms','Back','Wrist','Wrist2','Hands','Primary','Secondary','Finger','Finger2','Chest','Legs','Feet','Waist','Ammo'];
const CLASS_NAMES=['WAR','CLR','PAL','RNG','SHD','DRU','MNK','BRD','ROG','SHM','NEC','WIZ','MAG','ENC','BST','BER'];

function decodeSlots(m){{return SLOT_NAMES.filter((_,i)=>m&(1<<i)).join(', ')||'None';}}
function decodeClasses(m){{return m===65535?'ALL':CLASS_NAMES.filter((_,i)=>m&(1<<i)).join(' ')||'None';}}

function itemColor(d){{
  if(d.nodrop&&d.reqlevel>=45)return'#ff8000';
  if(d.nodrop&&(d.hp>100||d.ac>30||d.mana>100))return'#a335ee';
  if(d.nodrop||d.magic)return'#60a5fa';
  if(d.reqlevel>0||d.reclevel>0)return'#4ade80';
  return'#d0d0d0';
}}

function renderItemCard(el,d){{
  const color=itemColor(d);
  const flags=[
    d.magic?'<span class="item-flag" style="color:#60a5fa;border:1px solid #60a5fa">MAGIC</span>':'',
    d.nodrop?'<span class="item-flag" style="color:#f87171;border:1px solid #f87171">NO DROP</span>':'',
    d.norent?'<span class="item-flag" style="color:#fbbf24;border:1px solid #fbbf24">NO RENT</span>':'',
    d.attuneable?'<span class="item-flag" style="color:#a78bfa;border:1px solid #a78bfa">ATTUNABLE</span>':'',
  ].filter(Boolean).join(' ');
  const stat=(k,v,c)=>v?'<span>'+k+': <strong style="color:'+(c||'var(--text)')+'">'+( v>0?'+'+v:v)+'</strong></span>':'';
  const stats=[['STR',d.astr],['STA',d.asta],['AGI',d.aagi],['DEX',d.adex],['INT',d.aint],['WIS',d.awis],['CHA',d.acha]].filter(([,v])=>v).map(([k,v])=>stat(k,v)).join(' ');
  const heroics=[['H.STR',d.heroic_str],['H.STA',d.heroic_sta],['H.AGI',d.heroic_agi],['H.DEX',d.heroic_dex],['H.INT',d.heroic_int],['H.WIS',d.heroic_wis],['H.CHA',d.heroic_cha]].filter(([,v])=>v>0).map(([k,v])=>stat(k,v,'#a78bfa')).join(' ');
  const resists=[['CR',d.cr],['FR',d.fr],['MR',d.mr],['DR',d.dr],['PR',d.pr]].filter(([,v])=>v).map(([k,v])=>stat(k,v,'#22d3a0')).join(' ');
  el.innerHTML=`
    <button onclick="closeItemModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer;padding:0;line-height:1">&#10005;</button>
    <div style="font-size:1.05rem;font-weight:700;color:${{color}};margin-bottom:.25rem">${{d.name}}</div>
    ${{d.lore&&d.lore!==d.name?`<div style="font-size:.78rem;font-style:italic;color:var(--text3);margin-bottom:.5rem">${{d.lore}}</div>`:''}}
    <div style="margin-bottom:.65rem;display:flex;gap:.35rem;flex-wrap:wrap">${{flags||'<span style="color:var(--text3);font-size:.78rem">No flags</span>'}}</div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:.3rem .8rem;font-size:.82rem;margin-bottom:.65rem;color:var(--text2)">
      <div>Slot: <strong style="color:var(--text)">${{decodeSlots(d.slots)}}</strong></div>
      <div>Classes: <strong style="color:var(--text)">${{decodeClasses(d.classes)}}</strong></div>
      ${{d.reqlevel>0?`<div>Req Level: <strong style="color:var(--text)">${{d.reqlevel}}</strong></div>`:''}}
      ${{d.reclevel>0?`<div>Rec Level: <strong style="color:var(--text)">${{d.reclevel}}</strong></div>`:''}}
      ${{d.ac>0?`<div>AC: <strong style="color:var(--text)">${{d.ac}}</strong></div>`:''}}
      ${{d.damage>0?`<div>Dmg/Dly: <strong style="color:var(--text)">${{d.damage}}/${{d.delay}}</strong></div>`:''}}
      ${{d.range>0?`<div>Range: <strong style="color:var(--text)">${{d.range}}</strong></div>`:''}}
      ${{d.haste>0?`<div>Haste: <strong style="color:#fbbf24">+${{d.haste}}%</strong></div>`:''}}
      <div style="color:var(--text3)">Wt: ${{(d.weight/10).toFixed(1)}} &nbsp; ID: ${{d.id}}</div>
    </div>
    ${{(d.hp>0||d.mana>0||d.endur>0)?`<div style="font-size:.82rem;margin-bottom:.5rem;display:flex;gap:.8rem">${{d.hp>0?'<span>HP: <strong style=\"color:#4ade80\">+'+d.hp+'</strong></span>':''}}${{d.mana>0?'<span>Mana: <strong style=\"color:#60a5fa\">+'+d.mana+'</strong></span>':''}}${{d.endur>0?'<span>End: <strong style=\"color:#fbbf24\">+'+d.endur+'</strong></span>':''}}</div>`:''}}
    ${{stats?`<div style="font-size:.82rem;margin-bottom:.45rem;display:flex;gap:.6rem;flex-wrap:wrap">${{stats}}</div>`:''}}
    ${{heroics?`<div style="font-size:.82rem;margin-bottom:.45rem;display:flex;gap:.6rem;flex-wrap:wrap">${{heroics}}</div>`:''}}
    ${{resists?`<div style="font-size:.82rem;display:flex;gap:.6rem;flex-wrap:wrap">${{resists}}</div>`:''}}
    ${{(()=>{{
      const nm=d.name||'';
      if(!/^Spell:/i.test(nm)) return '';
      const sname=nm.replace(/^Spell:\\s*/i,'');
      const btnId='slb-'+d.id;
      // Use data-sname attribute to avoid JS string escaping in onclick
      return '<div style="margin-top:.65rem;padding-top:.6rem;border-top:1px solid var(--border);'
        +'display:flex;align-items:center;gap:.5rem;flex-wrap:wrap">'
        +'<button data-sname="'+sname.replace(/"/g,'&quot;')+'" data-bid="'+btnId+'" '
        +'onclick="const b=this;lookupSpellByName(b.dataset.sname,b.dataset.bid)" '
        +'style="background:var(--accent-bg);color:#a78bfa;border:1px solid #a78bfa55;'
        +'border-radius:5px;padding:.3rem .8rem;font-size:.8rem;cursor:pointer;font-weight:600">'
        +'&#10024; View Spell: '+sname+'</button>'
        +'<span id="'+btnId+'-status" style="font-size:.75rem;color:var(--text3)"></span>'
        +'</div>';
    }})()}}
  `;
}}

function lookupItem(itemId){{
  const modal=document.getElementById('item-modal');
  const el=document.getElementById('item-card-content');
  modal.classList.add('open');
  if(_itemCache[itemId]){{renderItemCard(el,_itemCache[itemId]);return;}}
  el.innerHTML='<button onclick="closeItemModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button><div style="text-align:center;padding:2.5rem;color:var(--text3)">Loading item #'+itemId+'...</div>';
  fetch('/spire/api/v1/item/'+itemId)
    .then(r=>{{if(!r.ok)throw new Error(r.status);return r.json();}})
    .then(d=>{{_itemCache[itemId]=d;renderItemCard(el,d);}})
    .catch(()=>{{el.innerHTML='<button onclick="closeItemModal()" style="position:absolute;top:.6rem;right:.8rem;background:none;border:none;color:var(--text3);font-size:1.2rem;cursor:pointer">&#10005;</button><div style="color:#f87171;padding:1.5rem;text-align:center">Failed to load item</div>';}});
}}

function closeItemModal(){{
  document.getElementById('item-modal').classList.remove('open');
}}

document.addEventListener('keydown',e=>{{if(e.key==='Escape'){{closeInvModal();closeItemModal();closeSpellModal();}}}});
window._allCmds = {cmds_json};
initTrendsExtra();
// Auto-refresh disabled
</script>
<div class="item-modal-overlay" id="inv-modal" onclick="if(event.target===this)closeInvModal()">
  <div class="item-card" id="inv-card-content" style="max-width:540px">
    <div style="text-align:center;padding:2rem;color:var(--text3)">Loading...</div>
  </div>
</div>
<div class="item-modal-overlay" id="spell-modal" onclick="if(event.target===this)closeSpellModal()">
  <div class="item-card" id="spell-card-content" style="max-width:480px">
    <div style="text-align:center;padding:2rem;color:var(--text3)">Loading...</div>
  </div>
</div>
<div class="item-modal-overlay" id="item-modal"  onclick="if(event.target===this)closeItemModal()">
  <div class="item-card" id="item-card-content">
    <div style="text-align:center;padding:2rem;color:var(--text3)">Loading...</div>
  </div>
</div>
</body>
</html>"""

with open(OUT_HTML,"w",encoding="utf-8") as f:
    f.write(html)


print(f"Excel: {OUT_XLSX}")
print(f"HTML:  {OUT_HTML}")
print(f"New since last report: {new_label}")
